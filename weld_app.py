"""
Weld Acquisition Web Interface
================================
Flask web app for managing weld experiments with NI-9201 DAQ.

Continuous acquisition with live threshold detection and real-time plotting.
- Starts continuous DAQ recording when "Start WELD" is pressed
- Keeps a rolling 2s pre-buffer; when threshold is crossed, that becomes the start
- Streams 1 datapoint/sec to the frontend for live dual-axis plot
- When signal drops below threshold + post_time, stops and saves HDF5
- Marks experiment as welded in the Excel database

Usage:
  conda activate distortion
  python weld_app.py
  # Open http://localhost:5000 in your browser
"""

import os
import sys
import json
import threading
import time
import collections
from datetime import datetime

import numpy as np
import openpyxl
import h5py
from flask import Flask, render_template, request, jsonify

# ---------------------------------------------------------------------------
# Try to import NI DAQ — allow running without hardware for UI development
# ---------------------------------------------------------------------------
try:
    import nidaqmx
    from nidaqmx.constants import AcquisitionType, TerminalConfiguration
    DAQ_AVAILABLE = True
except ImportError:
    DAQ_AVAILABLE = False

DEFAULT_DEVICE = "cDAQ4Mod1"
DEFAULT_RATE = 50_000

app = Flask(__name__)

# ---------------------------------------------------------------------------
# Global state
# ---------------------------------------------------------------------------
db_state = {
    "filepath": None,
    "rows": {},
    "loaded": False,
    "columns": [],
    "header_cols": [],
}

acq_state = {
    "running": False,
    "status": "idle",       # idle, waiting, welding, saving, done, error
    "message": "",
    "current_id": None,
    "output_file": None,
    "elapsed_s": 0,
}

# Plot data buffer — holds downsampled points (1/sec) for the frontend
# Each entry: {"t": seconds, "ch0": volts, "ch1": amps}
plot_buffer = []
plot_lock = threading.Lock()

acq_lock = threading.Lock()
abort_event = threading.Event()

# ---------------------------------------------------------------------------
# Acquisition config
# ---------------------------------------------------------------------------
ACQ_CONFIG = {
    "device": DEFAULT_DEVICE,
    "channels": [0, 1],       # ai0 = voltage, ai1 = current
    "rate": DEFAULT_RATE,
    "max_duration": 300.0,
    "threshold": 0.2,         # 10A / 50 A/V = 0.2V at DAQ input
    "trigger_channel": 1,     # trigger on current channel (ai1)
    "pre_time": 2.0,
    "post_time": 2.0,
    "output_dir": "measurements",
}

# ---------------------------------------------------------------------------
# Scaling factors: DAQ voltage → real units
# ---------------------------------------------------------------------------
# Voltage channel (ai0): 10:1 divider → 1V at DAQ = 10V real
VOLTAGE_SCALE = 10.0      # V_real = V_daq * 10
# Current channel (ai1): 10V at DAQ = 500A → 50 A/V
CURRENT_SCALE = 50.0      # A_real = V_daq * 50


# ---------------------------------------------------------------------------
# Database functions (unchanged)
# ---------------------------------------------------------------------------
def load_database(filepath):
    print(f"Loading database: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    headers = []
    header_cols = []
    rows = {}
    scan_max_col = 100
    row_idx = 0

    for row_cells in ws.iter_rows(min_col=1, max_col=scan_max_col, values_only=True):
        row_idx += 1
        if row_idx == 1:
            empty_streak = 0
            for col_num, val in enumerate(row_cells, start=1):
                if val is None:
                    empty_streak += 1
                    if empty_streak >= 5 and headers:
                        break
                    continue
                empty_streak = 0
                sval = str(val).strip()
                if not sval or sval.startswith("Spalte"):
                    continue
                headers.append(sval)
                header_cols.append(col_num)
            if not headers:
                wb.close()
                raise ValueError("No column headers found in row 1")
            continue

        if not row_cells or all(v is None for v in row_cells):
            continue

        row_data = {}
        for header, col_num in zip(headers, header_cols):
            idx = col_num - 1
            row_data[header] = row_cells[idx] if idx < len(row_cells) else None

        exp_id = row_data.get("ID")
        if exp_id is not None:
            row_data["_row_idx"] = row_idx
            rows[str(exp_id).strip().upper()] = row_data

    wb.close()

    db_state["filepath"] = filepath
    db_state["rows"] = rows
    db_state["loaded"] = True
    db_state["columns"] = headers
    db_state["header_cols"] = header_cols

    print(f"  Headers ({len(headers)}): {headers}")
    print(f"  Loaded {len(rows)} experiments")
    return len(rows)


def mark_as_welded(exp_id, h5_file=None):
    filepath = db_state["filepath"]
    if not filepath:
        return

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    row_data = db_state["rows"].get(exp_id)
    if not row_data:
        wb.close()
        return

    row_idx = row_data["_row_idx"]
    headers = db_state["columns"]
    header_cols = db_state.get("header_cols") or list(range(1, len(headers) + 1))

    welded_col = None
    h5file_col = None
    for i, h in enumerate(headers):
        if h == "WELDED":
            welded_col = header_cols[i]
        if h == "H5FILE":
            h5file_col = header_cols[i]

    if welded_col is None:
        welded_col = max(header_cols, default=0) + 1
        ws.cell(row=1, column=welded_col, value="WELDED")
        headers.append("WELDED")
        header_cols.append(welded_col)

    if h5file_col is None:
        h5file_col = max(header_cols, default=0) + 1
        ws.cell(row=1, column=h5file_col, value="H5FILE")
        headers.append("H5FILE")
        header_cols.append(h5file_col)

    ws.cell(row=row_idx, column=welded_col, value=datetime.now().isoformat())
    if h5_file:
        ws.cell(row=row_idx, column=h5file_col, value=os.path.basename(h5_file))

    wb.save(filepath)
    wb.close()

    row_data["WELDED"] = datetime.now().isoformat()
    if h5_file:
        row_data["H5FILE"] = os.path.basename(h5_file)
    db_state["columns"] = headers
    db_state["header_cols"] = header_cols


def unmark_welded(exp_id):
    """Remove WELDED and H5FILE marks from a row (undo a completed weld)."""
    filepath = db_state["filepath"]
    if not filepath:
        return False, "No database loaded"

    row_data = db_state["rows"].get(exp_id)
    if not row_data:
        return False, f"ID '{exp_id}' not found"

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    row_idx = row_data["_row_idx"]
    headers = db_state["columns"]
    header_cols = db_state.get("header_cols") or list(range(1, len(headers) + 1))

    welded_col = next((header_cols[i] for i, h in enumerate(headers) if h == "WELDED"), None)
    h5file_col = next((header_cols[i] for i, h in enumerate(headers) if h == "H5FILE"), None)

    if welded_col is None:
        wb.close()
        return False, "No WELDED column found — nothing to undo"

    ws.cell(row=row_idx, column=welded_col, value=_excel_write_value(None))
    if h5file_col:
        ws.cell(row=row_idx, column=h5file_col, value=_excel_write_value(None))

    wb.save(filepath)
    wb.close()

    row_data.pop("WELDED", None)
    row_data.pop("H5FILE", None)

    print(f"Unmarked {exp_id} as welded")
    return True, f"ID {exp_id} zurückgesetzt"


EDITABLE_FIELDS = {"WFS [m/min]", "WS [m/min]", "PATH ID", "SERIES", "NUMBER", "COMMENT"}
UPPERCASE_STRING_FIELDS = {"PATH ID"}
STRING_FIELDS = {"PATH ID", "COMMENT"}


def _excel_write_value(val):
    """openpyxl leaves existing content when value=None; use '' to clear a cell."""
    return "" if val is None else val


def _resolve_exp_row(exp_id):
    exp_id = str(exp_id).strip().upper()
    row = db_state["rows"].get(exp_id)
    if row is not None:
        return exp_id, row
    for key, row in db_state["rows"].items():
        if str(key).strip().upper() == exp_id:
            return key, row
    return exp_id, None


def _resolve_column(field):
    field = str(field).strip()
    headers = db_state["columns"]
    header_cols = db_state.get("header_cols") or list(range(1, len(headers) + 1))
    if field in headers:
        idx = headers.index(field)
        return header_cols[idx], headers[idx]
    field_upper = field.upper()
    for i, h in enumerate(headers):
        if str(h).strip().upper() == field_upper:
            return header_cols[i], h
    return None, None


def update_row_field(exp_id, field, value):
    """Update a single editable column in the Excel database."""
    field = str(field).strip()
    if field not in EDITABLE_FIELDS and field.upper() not in {f.upper() for f in EDITABLE_FIELDS}:
        return False, f"Field '{field}' is not editable", None

    filepath = db_state["filepath"]
    if not filepath:
        return False, "No database loaded", None

    _, row_data = _resolve_exp_row(exp_id)
    if row_data is None:
        return False, f"ID '{exp_id}' not found", None

    if value is None or (isinstance(value, str) and not value.strip()):
        val = None
    elif isinstance(value, str) and value.strip().lower() in ("none", "null"):
        val = None
    elif field in STRING_FIELDS or field.upper() in {f.upper() for f in STRING_FIELDS}:
        val = str(value).strip()
        if field in UPPERCASE_STRING_FIELDS or field.upper() == "PATH ID":
            val = val.upper()
    else:
        try:
            val = float(value)
            if val == int(val):
                val = int(val)
        except (TypeError, ValueError):
            return False, "Invalid number", None

    excel_col, col_name = _resolve_column(field)
    if excel_col is None:
        return False, f"Column '{field}' not found in database", None

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        ws.cell(row=row_data["_row_idx"], column=excel_col, value=_excel_write_value(val))
        wb.save(filepath)
        wb.close()
    except Exception as e:
        return False, f"Excel save failed: {e}", None

    row_data[col_name] = val
    print(f"Updated {exp_id} {col_name} = {val!r}")
    return True, f"{col_name} saved", col_name


# ---------------------------------------------------------------------------
# Continuous acquisition with live threshold detection
# ---------------------------------------------------------------------------
def run_acquisition(exp_id, row_data):
    """
    Background thread: continuous NI-DAQmx acquisition.

    States:
      waiting  — acquiring, threshold not yet crossed
      welding  — threshold crossed, recording weld data
      saving   — signal dropped + post_time elapsed, saving HDF5
      done     — complete
    """
    cfg = ACQ_CONFIG
    rate = cfg["rate"]
    channels = cfg["channels"]
    num_channels = len(channels)
    threshold = cfg["threshold"]
    trigger_ch = cfg["trigger_channel"]  # index into channels list
    pre_time = cfg["pre_time"]
    post_time = cfg["post_time"]
    max_duration = cfg["max_duration"]

    pre_samples = int(pre_time * rate)
    post_samples = int(post_time * rate)

    # Chunk size: read 0.1s of data at a time
    chunk_samples = max(1, int(rate * 0.1))

    os.makedirs(cfg["output_dir"], exist_ok=True)
    ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    outfile = os.path.join(cfg["output_dir"], f"{exp_id}_{ts_str}.h5")

    abort_event.clear()

    with acq_lock:
        acq_state["running"] = True
        acq_state["status"] = "waiting"
        acq_state["message"] = "Acquiring... waiting for weld signal"
        acq_state["current_id"] = exp_id
        acq_state["output_file"] = outfile
        acq_state["elapsed_s"] = 0

    with plot_lock:
        plot_buffer.clear()

    try:
        if not DAQ_AVAILABLE:
            _simulate_acquisition(exp_id, row_data, outfile)
            return

        _run_real_acquisition(
            exp_id, row_data, outfile, cfg, rate, channels, num_channels,
            threshold, trigger_ch, pre_samples, post_samples,
            chunk_samples, max_duration,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        with acq_lock:
            acq_state["status"] = "error"
            acq_state["message"] = f"Error: {str(e)}"
            acq_state["running"] = False


def _run_real_acquisition(exp_id, row_data, outfile, cfg, rate, channels,
                          num_channels, threshold, trigger_ch,
                          pre_samples, post_samples, chunk_samples, max_duration):
    """Real NI-DAQmx continuous acquisition with live threshold detection."""
    device = cfg["device"]
    chan_string = ", ".join(f"{device}/ai{ch}" for ch in channels)

    # All collected data (list of numpy arrays per channel)
    all_data = [[] for _ in range(num_channels)]
    # Rolling pre-buffer (circular) for pre-trigger data
    pre_buf = [collections.deque(maxlen=pre_samples) for _ in range(num_channels)]

    threshold_crossed = False
    weld_start_idx = 0
    below_threshold_since = None
    total_samples = 0
    start_time = time.time()
    last_plot_time = 0

    with nidaqmx.Task() as task:
        task.ai_channels.add_ai_voltage_chan(
            chan_string,
            min_val=-10.0,
            max_val=10.0,
            terminal_config=TerminalConfiguration.RSE,
        )
        task.timing.cfg_samp_clk_timing(
            rate=rate,
            sample_mode=AcquisitionType.CONTINUOUS,
            samps_per_chan=chunk_samples * 10,  # buffer size
        )
        task.start()

        print(f"Continuous acquisition started: {chan_string} @ {rate} S/s")

        while True:
            if abort_event.is_set():
                print("Acquisition aborted by user")
                with acq_lock:
                    acq_state["status"] = "idle"
                    acq_state["message"] = "Aborted."
                    acq_state["running"] = False
                task.stop()
                start_monitor()
                return

            elapsed = time.time() - start_time
            if elapsed > max_duration:
                print("Max duration reached")
                with acq_lock:
                    acq_state["message"] = "Max duration reached without weld completion"
                break

            # Read a chunk
            try:
                data = task.read(
                    number_of_samples_per_channel=chunk_samples,
                    timeout=5.0,
                )
            except Exception as e:
                print(f"Read error: {e}")
                break

            # Convert to numpy — handle single vs multi channel
            data = np.array(data)
            if data.ndim == 1:
                data = data.reshape(1, -1)

            chunk_len = data.shape[1]
            total_samples += chunk_len

            with acq_lock:
                acq_state["elapsed_s"] = elapsed

            # --- Downsampled plot point (1/sec) ---
            now = time.time()
            if now - last_plot_time >= 1.0:
                last_plot_time = now
                # Average of this chunk for display
                pt = {"t": round(elapsed, 1)}
                pt["ch0"] = round(float(np.mean(data[0])) * VOLTAGE_SCALE, 4)
                if num_channels > 1:
                    pt["ch1"] = round(float(np.mean(data[1])) * CURRENT_SCALE, 4)
                else:
                    pt["ch1"] = 0.0
                with plot_lock:
                    plot_buffer.append(pt)

            # --- Threshold logic ---
            trigger_data = data[trigger_ch]

            if not threshold_crossed:
                # Fill pre-buffer
                for ch in range(num_channels):
                    pre_buf[ch].extend(data[ch])

                # Robust start detection: require chunk MEAN above threshold
                # (single-sample spikes from wire inching, brief ignition
                # attempts etc. won't trigger a false start anymore).
                chunk_mean = float(np.mean(trigger_data))
                if chunk_mean > threshold:
                    threshold_crossed = True

                    # Flush pre-buffer (already contains this chunk's data) into all_data
                    for ch in range(num_channels):
                        pre_arr = np.array(pre_buf[ch])
                        all_data[ch].append(pre_arr)
                    pre_buf = None

                    with acq_lock:
                        acq_state["status"] = "welding"
                        acq_state["message"] = f"Weld detected! Recording... ({elapsed:.0f}s)"

                    print(f"Threshold crossed at {elapsed:.2f}s (chunk mean={chunk_mean:.4f}V)")

            else:
                # We're in welding mode — accumulate all data
                for ch in range(num_channels):
                    all_data[ch].append(data[ch].copy())

                # Check if signal dropped below threshold.
                # Use chunk mean (not np.any) so that noise spikes after
                # weld end don't keep resetting the post-time timer.
                chunk_mean = float(np.mean(trigger_data))
                if chunk_mean <= threshold:
                    if below_threshold_since is None:
                        below_threshold_since = time.time()
                        print(f"Signal below threshold (mean={chunk_mean:.4f}V) — post-time countdown started")
                    elif time.time() - below_threshold_since >= cfg["post_time"]:
                        print(f"Signal below threshold for {cfg['post_time']}s — stopping")
                        with acq_lock:
                            acq_state["status"] = "saving"
                            acq_state["message"] = "Weld complete. Saving data..."
                        break
                else:
                    if below_threshold_since is not None:
                        print(f"Signal back above threshold (mean={chunk_mean:.4f}V) — post-time reset")
                    below_threshold_since = None

                with acq_lock:
                    acq_state["message"] = f"Recording weld... ({elapsed:.0f}s)"

        task.stop()

    # --- Save data ---
    with acq_lock:
        acq_state["status"] = "saving"
        acq_state["message"] = "Saving HDF5..."

    # Concatenate all chunks per channel
    final_data = np.zeros((num_channels, 0))
    if any(len(ch_list) > 0 for ch_list in all_data):
        ch_arrays = []
        for ch in range(num_channels):
            if all_data[ch]:
                ch_arrays.append(np.concatenate(all_data[ch]))
            else:
                ch_arrays.append(np.array([]))
        # Ensure all same length
        min_len = min(len(a) for a in ch_arrays)
        if min_len > 0:
            final_data = np.array([a[:min_len] for a in ch_arrays])

    num_samples = final_data.shape[1] if final_data.ndim == 2 else 0
    path_id = row_data.get("PATH ID", "XX")
    wfs = row_data.get("WFS [m/min]", 0)

    if num_samples > 0:
        time_axis = np.arange(num_samples) / rate
        timestamps = np.array([
            datetime.now().timestamp() - (num_samples - i) / rate
            for i in range(num_samples)
        ])

        with h5py.File(outfile, "w") as f:
            # --- Acquisition metadata ---
            f.attrs["device"] = f"NI cDAQ-9171 + NI-9201 ({cfg['device']})"
            f.attrs["timestamp"] = datetime.now().isoformat()
            f.attrs["sample_rate_per_channel_hz"] = rate
            f.attrs["aggregate_rate_hz"] = rate * num_channels
            f.attrs["num_channels"] = num_channels
            f.attrs["channel_indices"] = channels
            f.attrs["input_range"] = "BIP10VOLTS"
            f.attrs["duration_s"] = num_samples / rate
            f.attrs["threshold_daq_v"] = threshold
            f.attrs["threshold_real_A"] = threshold * CURRENT_SCALE
            f.attrs["voltage_scale"] = VOLTAGE_SCALE
            f.attrs["current_scale"] = CURRENT_SCALE
            f.attrs["pre_time_s"] = cfg["pre_time"]
            f.attrs["post_time_s"] = cfg["post_time"]

            # --- Experiment parameters (all columns from Excel) ---
            for key, val in row_data.items():
                if key.startswith("_"):
                    continue  # skip internal keys like _row_idx
                attr_key = key.replace(" ", "_").replace("[", "").replace("]", "").replace("/", "_per_")
                if val is None:
                    f.attrs[f"exp_{attr_key}"] = ""
                elif isinstance(val, (int, float)):
                    f.attrs[f"exp_{attr_key}"] = val
                else:
                    f.attrs[f"exp_{attr_key}"] = str(val)

            f.create_dataset("time_s", data=time_axis,
                             compression="gzip", chunks=(min(100_000, num_samples),))
            f.create_dataset("timestamps_unix", data=timestamps,
                             compression="gzip", chunks=(min(100_000, num_samples),))

            # Store scaled real-world values
            # ai0 = voltage: DAQ * 10 (10:1 divider)
            # ai1 = current: DAQ * 50 (10V = 500A)
            scales = {0: VOLTAGE_SCALE, 1: CURRENT_SCALE}
            units = {0: "V", 1: "A"}
            labels = {0: "voltage", 1: "current"}

            for ch_idx, ch_num in enumerate(channels):
                scale = scales.get(ch_num, 1.0)
                unit = units.get(ch_num, "volts")
                label = labels.get(ch_num, f"channel_{ch_num}")
                ds = f.create_dataset(
                    f"channel_{ch_num}", data=final_data[ch_idx] * scale,
                    compression="gzip", compression_opts=4,
                    chunks=(min(100_000, num_samples),),
                )
                ds.attrs["units"] = unit
                ds.attrs["label"] = label
                ds.attrs["channel_number"] = ch_num
                ds.attrs["physical_channel"] = f"{cfg['device']}/ai{ch_num}"
                ds.attrs["sample_rate_hz"] = rate
                ds.attrs["scale_factor"] = scale

        file_size = os.path.getsize(outfile)
        print(f"Saved {outfile}: {num_samples:,} samples/ch, "
              f"{num_samples/rate:.2f}s, {file_size/1024/1024:.1f} MB")

    mark_as_welded(exp_id, outfile)

    with acq_lock:
        acq_state["status"] = "done"
        acq_state["message"] = (
            f"Weld complete! {num_samples/rate:.1f}s recorded. "
            f"Saved {os.path.basename(outfile)}"
        )
        acq_state["running"] = False


def _simulate_acquisition(exp_id, row_data, outfile):
    """Simulate continuous acquisition for UI testing without NI hardware."""
    cfg = ACQ_CONFIG
    rate = cfg["rate"]
    channels = cfg["channels"]
    threshold = cfg["threshold"]
    path_id = row_data.get("PATH ID", "XX")
    wfs = row_data.get("WFS [m/min]", 0)

    start_time = time.time()
    sim_data_ch0 = []
    sim_data_ch1 = []

    # Simulate: 3s idle, then 8s weld signal, then drop
    weld_start = 3.0
    weld_end = 11.0
    total_sim = weld_end + cfg["post_time"] + 1.0

    t = 0
    while t < total_sim:
        if abort_event.is_set():
            print("Simulation aborted by user")
            with acq_lock:
                acq_state["status"] = "idle"
                acq_state["message"] = "Aborted."
                acq_state["running"] = False
            start_monitor()
            return
        time.sleep(0.5)
        t = time.time() - start_time

        # Generate simulated values (in real units: V and A)
        if weld_start <= t <= weld_end:
            v0 = 22.0 + np.random.randn() * 1.5    # ~22V welding voltage
            v1 = 150.0 + np.random.randn() * 10.0   # ~150A welding current
            sim_data_ch0.append(v0)
            sim_data_ch1.append(v1)
        else:
            v0 = np.random.randn() * 0.5
            v1 = np.random.randn() * 2.0
            sim_data_ch0.append(v0)
            sim_data_ch1.append(v1)

        # Update plot buffer (already in real units)
        pt = {"t": round(t, 1), "ch0": round(v0, 4), "ch1": round(v1, 4)}
        with plot_lock:
            plot_buffer.append(pt)

        # Update state
        with acq_lock:
            acq_state["elapsed_s"] = t
            if t < weld_start:
                acq_state["status"] = "waiting"
                acq_state["message"] = f"[SIM] Waiting for weld signal... ({t:.0f}s)"
            elif t <= weld_end:
                acq_state["status"] = "welding"
                acq_state["message"] = f"[SIM] Recording weld... ({t:.0f}s)"
            elif t <= weld_end + cfg["post_time"]:
                acq_state["status"] = "welding"
                acq_state["message"] = f"[SIM] Signal dropped, post-time... ({t:.0f}s)"
            else:
                break

    with acq_lock:
        acq_state["status"] = "saving"
        acq_state["message"] = "[SIM] Saving data..."

    # Save simulated HDF5
    n = len(sim_data_ch0)
    time_axis = np.arange(n) * 0.5  # 0.5s between sim points
    timestamps = np.array([start_time + i * 0.5 for i in range(n)])

    with h5py.File(outfile, "w") as f:
        f.attrs["device"] = "SIMULATED"
        f.attrs["timestamp"] = datetime.now().isoformat()
        f.attrs["sample_rate_per_channel_hz"] = 2  # sim rate
        f.attrs["experiment_id"] = exp_id
        f.attrs["path_id"] = str(path_id)
        f.attrs["wfs_m_per_min"] = float(wfs) if wfs else 0
        f.attrs["duration_s"] = n * 0.5
        f.create_dataset("time_s", data=time_axis, compression="gzip")
        f.create_dataset("timestamps_unix", data=timestamps, compression="gzip")
        f.create_dataset("channel_0", data=np.array(sim_data_ch0), compression="gzip")
        f.create_dataset("channel_1", data=np.array(sim_data_ch1), compression="gzip")

    mark_as_welded(exp_id, outfile)

    with acq_lock:
        acq_state["status"] = "done"
        acq_state["message"] = f"[SIM] Weld complete! Saved {os.path.basename(outfile)}"
        acq_state["running"] = False


# ---------------------------------------------------------------------------
# Live monitor — reads DAQ values continuously in background
# ---------------------------------------------------------------------------
monitor_state = {
    "active": False,
    "voltage": 0.0,
    "current": 0.0,
}
monitor_stop_event = threading.Event()


def monitor_loop():
    """Background thread: continuously read single samples from DAQ."""
    cfg = ACQ_CONFIG
    device = cfg["device"]
    channels = cfg["channels"]

    if DAQ_AVAILABLE:
        chan_string = ", ".join(f"{device}/ai{ch}" for ch in channels)
        task = nidaqmx.Task()
        task.ai_channels.add_ai_voltage_chan(
            chan_string,
            min_val=-10.0, max_val=10.0,
            terminal_config=TerminalConfiguration.RSE,
        )
        # Single sample on-demand reads — no timing needed

    print("Monitor started")
    monitor_state["active"] = True

    try:
        while not monitor_stop_event.is_set():
            if DAQ_AVAILABLE:
                try:
                    vals = task.read()
                    if isinstance(vals, list):
                        v_raw = vals[0]
                        i_raw = vals[1] if len(vals) > 1 else 0.0
                    else:
                        v_raw = vals
                        i_raw = 0.0
                    monitor_state["voltage"] = round(v_raw * VOLTAGE_SCALE, 3)
                    monitor_state["current"] = round(i_raw * CURRENT_SCALE, 3)
                except Exception as e:
                    monitor_state["voltage"] = 0.0
                    monitor_state["current"] = 0.0
            else:
                # Simulation: small random noise
                monitor_state["voltage"] = round(np.random.randn() * 0.5, 3)
                monitor_state["current"] = round(np.random.randn() * 2.0, 3)

            time.sleep(0.5)
    finally:
        if DAQ_AVAILABLE:
            task.close()
        monitor_state["active"] = False
        print("Monitor stopped")


def start_monitor():
    """Start the monitor if not already running."""
    if monitor_state["active"]:
        return
    monitor_stop_event.clear()
    t = threading.Thread(target=monitor_loop, daemon=True)
    t.start()


def stop_monitor():
    """Stop the monitor (e.g. when weld acquisition starts)."""
    monitor_stop_event.set()
    # Give it time to release the DAQ task
    time.sleep(0.3)


# ---------------------------------------------------------------------------
# H5 replay — load recorded weld data for the chart
# ---------------------------------------------------------------------------
def resolve_h5_path(h5_name):
    """Resolve H5FILE basename to an absolute path under output_dir."""
    if not h5_name:
        return None
    if os.path.isabs(h5_name) and os.path.exists(h5_name):
        return h5_name
    base_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(base_dir, ACQ_CONFIG["output_dir"], os.path.basename(h5_name)),
        os.path.join(base_dir, os.path.basename(h5_name)),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return candidates[0]


def load_h5_plot_points(h5_path):
    """Downsample HDF5 weld data to ~1 point/sec for the frontend chart."""
    with h5py.File(h5_path, "r") as f:
        time_s = f["time_s"][:]
        ch0 = f["channel_0"][:] if "channel_0" in f else np.zeros_like(time_s)
        ch1 = f["channel_1"][:] if "channel_1" in f else np.zeros_like(time_s)

    if len(time_s) == 0:
        return []

    second_bins = np.floor(time_s).astype(int)
    points = []
    for sec in range(int(second_bins[-1]) + 1):
        mask = second_bins == sec
        if not np.any(mask):
            continue
        points.append({
            "t": round(float(sec), 1),
            "ch0": round(float(np.mean(ch0[mask])), 4),
            "ch1": round(float(np.mean(ch1[mask])), 4),
        })
    return points


# ---------------------------------------------------------------------------
# Flask routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/load_db", methods=["POST"])
def api_load_db():
    data = request.get_json(silent=True) or {}
    filepath = data.get("filepath", "").strip()

    if not filepath:
        return jsonify({"error": "No file path provided"}), 400

    if not os.path.isabs(filepath):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        candidates = [
            os.path.normpath(os.path.join(base_dir, filepath)),
            os.path.normpath(os.path.join(os.getcwd(), filepath)),
            filepath,
        ]
        for candidate in candidates:
            if os.path.exists(candidate):
                filepath = candidate
                break

    if not os.path.exists(filepath):
        return jsonify({"error": f"File not found: {filepath}"}), 404

    try:
        count = load_database(filepath)
        return jsonify({"ok": True, "count": count})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/lookup/<exp_id>")
def api_lookup(exp_id):
    if not db_state["loaded"]:
        return jsonify({"error": "No database loaded"}), 400

    exp_id = exp_id.strip().upper()
    row = db_state["rows"].get(exp_id)
    if row is None:
        for key, candidate in db_state["rows"].items():
            if str(key).strip().upper() == exp_id:
                row = candidate
                break

    if row is None:
        return jsonify({"error": f"ID '{exp_id}' not found in database"}), 404

    result = {}
    for key, val in row.items():
        if key.startswith("_"):
            continue
        # Ensure JSON-serializable
        if isinstance(val, str) and not val.strip():
            val = None
        elif isinstance(val, str) and val.strip().lower() in ("none", "null"):
            val = None
        if isinstance(val, (int, float, str, bool)) or val is None:
            result[key] = val
        else:
            result[key] = str(val)

    result["already_welded"] = row.get("WELDED") is not None
    return jsonify(result)


@app.route("/api/update_field", methods=["POST"])
def api_update_field():
    if not db_state["loaded"]:
        return jsonify({"error": "No database loaded"}), 400

    data = request.get_json(silent=True) or {}
    exp_id = data.get("id", "").strip().upper()
    field = data.get("field", "").strip()
    value = data.get("value") if "value" in data else None

    if not exp_id or not field:
        return jsonify({"error": "Missing id or field"}), 400

    ok, msg, col_name = update_row_field(exp_id, field, value)
    if ok:
        _, row_data = _resolve_exp_row(exp_id)
        stored = row_data.get(col_name) if row_data else None
        return jsonify({"ok": True, "message": msg, "field": col_name, "value": stored})
    return jsonify({"error": msg}), 400


@app.route("/api/start_weld", methods=["POST"])
def api_start_weld():
    data = request.get_json()
    exp_id = data.get("id", "").strip().upper()

    if not db_state["loaded"]:
        return jsonify({"error": "No database loaded"}), 400

    row = db_state["rows"].get(exp_id)
    if row is None:
        return jsonify({"error": f"ID '{exp_id}' not found"}), 404

    with acq_lock:
        if acq_state["running"]:
            return jsonify({"error": "Acquisition already in progress"}), 409

    # Stop monitor so it releases the DAQ device
    stop_monitor()

    thread = threading.Thread(target=run_acquisition, args=(exp_id, row), daemon=True)
    thread.start()

    return jsonify({"ok": True, "message": "Acquisition started"})


@app.route("/api/status")
def api_status():
    with acq_lock:
        return jsonify({
            "running": acq_state["running"],
            "status": acq_state["status"],
            "message": acq_state["message"],
            "current_id": acq_state["current_id"],
            "output_file": acq_state["output_file"],
            "elapsed_s": acq_state["elapsed_s"],
        })


@app.route("/api/plot_data")
def api_plot_data():
    """Return plot data points since a given index."""
    since = int(request.args.get("since", 0))
    with plot_lock:
        new_points = plot_buffer[since:]
        total = len(plot_buffer)
    return jsonify({"points": new_points, "total": total})


@app.route("/api/h5_plot/<exp_id>")
def api_h5_plot(exp_id):
    """Return downsampled voltage/current plot data from a recorded H5 file."""
    if not db_state["loaded"]:
        return jsonify({"error": "No database loaded"}), 400

    exp_id = exp_id.strip().upper()
    row = db_state["rows"].get(exp_id)
    if row is None:
        return jsonify({"error": f"ID '{exp_id}' not found in database"}), 404

    h5_name = row.get("H5FILE")
    if not h5_name:
        return jsonify({"error": "No H5 file recorded for this experiment"}), 404

    h5_path = resolve_h5_path(str(h5_name))
    if not os.path.exists(h5_path):
        return jsonify({"error": f"H5 file not found: {h5_name}"}), 404

    try:
        points = load_h5_plot_points(h5_path)
        return jsonify({
            "points": points,
            "total": len(points),
            "file": os.path.basename(h5_path),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/abort", methods=["POST"])
def api_abort():
    """Stop any running acquisition immediately, without saving or marking as welded."""
    abort_event.set()
    # Give the acquisition thread up to 1s to notice and exit cleanly
    deadline = time.time() + 1.0
    while time.time() < deadline:
        with acq_lock:
            if not acq_state["running"]:
                break
        time.sleep(0.05)
    # Force-clear state in case thread didn't exit yet
    with acq_lock:
        acq_state["running"] = False
        acq_state["status"] = "idle"
        acq_state["message"] = "Aborted."
        acq_state["current_id"] = None
        acq_state["output_file"] = None
        acq_state["elapsed_s"] = 0
    with plot_lock:
        plot_buffer.clear()
    return jsonify({"ok": True})


@app.route("/api/reset", methods=["POST"])
def api_reset():
    with acq_lock:
        acq_state["running"] = False
        acq_state["status"] = "idle"
        acq_state["message"] = ""
        acq_state["current_id"] = None
        acq_state["output_file"] = None
        acq_state["elapsed_s"] = 0
    with plot_lock:
        plot_buffer.clear()
    # Restart live monitor
    start_monitor()
    return jsonify({"ok": True})


@app.route("/api/unmark/<exp_id>", methods=["POST"])
def api_unmark(exp_id):
    """Remove the WELDED/H5FILE mark for a given experiment ID."""
    if not db_state["loaded"]:
        return jsonify({"error": "No database loaded"}), 400
    exp_id = exp_id.strip().upper()
    ok, msg = unmark_welded(exp_id)
    if ok:
        return jsonify({"ok": True, "message": msg})
    return jsonify({"error": msg}), 400


@app.route("/api/monitor")
def api_monitor():
    """Return current live DAQ readings (scaled)."""
    return jsonify({
        "active": monitor_state["active"],
        "voltage": monitor_state["voltage"],
        "current": monitor_state["current"],
    })


@app.route("/api/config", methods=["GET", "POST"])
def api_config():
    if request.method == "GET":
        return jsonify(ACQ_CONFIG)

    data = request.get_json()
    for key in ["device", "channels", "rate", "max_duration",
                "threshold", "trigger_channel", "pre_time", "post_time"]:
        if key in data:
            ACQ_CONFIG[key] = data[key]
    return jsonify({"ok": True, "config": ACQ_CONFIG})


if __name__ == "__main__":
    print("=" * 60)
    print("Weld Acquisition Web Interface")
    print("=" * 60)
    if not DAQ_AVAILABLE:
        print("WARNING: NI-DAQmx not available — running in SIMULATION mode")
    print("Open http://localhost:5000 in your browser")
    print("=" * 60)
    start_monitor()
    app.run(debug=True, host="0.0.0.0", port=5000, use_reloader=False)
