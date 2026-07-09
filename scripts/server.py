from pathlib import Path
import json
import os
import re
import subprocess
import sys
import threading
import time
from datetime import datetime

import h5py
import numpy as np
import pandas as pd
from flask import Flask, jsonify, redirect, render_template, request, send_from_directory, url_for
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = BASE_DIR / "templates"
DEFAULT_VERSUCHSUEBERSICHT = BASE_DIR / "data" / "Versuchsübersicht.xlsx"
DEFAULT_VERSUCHSUEBERSICHT_PATH = "data/Versuchsübersicht.xlsx"

app = Flask(__name__, template_folder=str(TEMPLATES_DIR))

versuchsuebersicht_path = None
versuchsuebersicht_data = None
_scan_document_cache = {}
_scan_jobs = {}
_scan_jobs_lock = threading.Lock()
_excel_workbook_lock = threading.Lock()
_weld_jobs = {}
_weld_jobs_lock = threading.Lock()
_weld_abort_events = {}

MENU_TABLE_COLUMNS = ["COUNT", "ID", "WELDED", "SCANNED", "SERIES", "NUMBER", "COMMENT"]
ANALYZING_TABLE_COLUMNS = ["COUNT", "ID", "WELDED", "SCANNED"]
MENU_FILTER_OPTIONS = ["WELDED", "SCANNED", "ID"]
WELDING_FORM_COLUMNS = ["PATH ID", "WFS [m/min]", "WS [m/min]", "SERIES", "NUMBER", "COMMENT"]
SCANSPEED_COLUMN = "SCANSPEED [mm/s]"
WELDING_PARAM_TABLE_COLUMNS = [
    ["SERIES", "NUMBER", "WFS [m/min]", "WS [m/min]", "WELDED", "H5FILE"],
    ["WIRE", "WIREDIA [mm]", "GAS", "GASFLOW [l/min]", "POWERSOURCE", "PROCESS"],
    ["RESOLUTION", "PROFILEFREQUENCY", SCANSPEED_COLUMN, "SCANDURATION", "SCANNED", "JSONFILE"],
]
WELDING_PARAM_READONLY_COLUMNS = {
    "WELDED",
    "H5FILE",
    "WIRE",
    "WIREDIA [mm]",
    "GAS",
    "GASFLOW [l/min]",
    "POWERSOURCE",
    "PROCESS",
    "RESOLUTION",
    "PROFILEFREQUENCY",
    SCANSPEED_COLUMN,
    "SCANDURATION",
    "SCANNED",
    "JSONFILE",
}
SCANNING_EDITABLE_COLUMNS = [
    "RESOLUTION",
    "PROFILEFREQUENCY",
    SCANSPEED_COLUMN,
    "SCANDURATION",
    "COMMENT",
]
SCANNING_EDITABLE_AFTER_SCAN_COLUMNS = ["COMMENT"]
SCANNING_PARAM_READONLY_COLUMNS = {
    name
    for name in (
        field_name
        for block in WELDING_PARAM_TABLE_COLUMNS
        for field_name in block
    )
    if name not in {"RESOLUTION", "PROFILEFREQUENCY", SCANSPEED_COLUMN, "SCANDURATION"}
}
PATH_ID_OPTIONS = ["SO", "PO", "SI", "ME", "LI"]
PATH_IMAGES_DIR = TEMPLATES_DIR / "path_images"
WELDING_DATA_DIR = BASE_DIR / "data" / "welding"
SCANNING_DATA_DIR = BASE_DIR / "data" / "scanning"
ANALYZING_DATA_DIR = BASE_DIR / "data" / "analyzing"
SCAN_SCRIPT_PATH = BASE_DIR / "scripts" / "run_scan.py"
WELD_SCRIPT_PATH = BASE_DIR / "scripts" / "run_weld.py"
WELDING_METADATA_COLUMNS = ["H5FILE"]
SCANNING_METADATA_COLUMNS = ["JSONFILE"]
ANALYZING_METADATA_COLUMNS = ["ANALYZESCAN", "ANALYZEWELD"]
SCAN_SETTINGS_COLUMNS = ["RESOLUTION", "PROFILEFREQUENCY", SCANSPEED_COLUMN, "SCANDURATION"]
SCANNING_PRESCAN_EXCEL_COLUMNS = SCAN_SETTINGS_COLUMNS
SCAN_DEFAULT_EXPOSURE_US = 100
SCAN_RESOLUTION_OPTIONS = [160, 320, 640, 1280]
SCAN_DURATION_AUTO_SECONDS = 1.0
EXCEL_INTEGER_COLUMNS = {
    "SERIES",
    "NUMBER",
    "RESOLUTION",
    "PROFILEFREQUENCY",
}
EXCEL_FLOAT_COLUMNS = {
    "WFS [m/min]",
    "WS [m/min]",
    "SCANDURATION",
    SCANSPEED_COLUMN,
}
EXCEL_NUMERIC_COLUMNS = EXCEL_INTEGER_COLUMNS | EXCEL_FLOAT_COLUMNS
SCAN_USE_DEMO = os.environ.get("SCAN_USE_DEMO", "").strip().lower() in {"1", "true", "yes"}
SCANNER_IP = os.environ.get("SCANNER_IP", "").strip() or None
WELD_USE_DEMO = os.environ.get("WELD_USE_DEMO", "1").strip().lower() in {"1", "true", "yes"}
WELD_DEVICE = os.environ.get("WELD_DEVICE", "cDAQ4Mod1").strip() or "cDAQ4Mod1"
WELD_RATE = float(os.environ.get("WELD_RATE", "50000"))
WELD_THRESHOLD_V = float(os.environ.get("WELD_THRESHOLD_V", "0.2"))
WELD_PRE_TIME_S = float(os.environ.get("WELD_PRE_TIME_S", "2.0"))
WELD_POST_TIME_S = float(os.environ.get("WELD_POST_TIME_S", "2.0"))
WELD_MAX_DURATION_S = float(os.environ.get("WELD_MAX_DURATION_S", "300.0"))
WELDING_GRAPH_MAX_POINTS = 2000
SCANNING_GRAPH_MAX_POINTS = 400
DEFAULT_TABLE_FILTER_COLUMN = "WELDED"
EXCEL_SUFFIXES = {".xlsx", ".xls", ".xlsm"}
MAX_ROWS = 1000
_DEFAULT_EXCEL_PLACEHOLDER_COLUMN = re.compile(r"^spalte\d+$", re.IGNORECASE)


def _find_id_column(columns):
    return _find_column(columns, "id")


def _find_column(columns, name):
    target = name.strip().lower()
    for column in columns:
        if str(column).strip().lower() == target:
            return str(column)
    return None


def _normalize_id(value):
    if value is None:
        return None

    text = str(value).strip().upper()
    if not text:
        return None

    return text


def _resolve_path(path_str):
    path = Path(path_str)
    if not path.is_absolute():
        path = BASE_DIR / path
    return path


def _build_id_index(rows, columns):
    id_column = _find_id_column(columns)
    if id_column is None:
        return {}

    index = {}
    for row in rows:
        normalized_id = _normalize_id(row.get(id_column))
        if normalized_id:
            index[normalized_id] = row
    return index


def _filter_rows_with_id(df):
    id_column = _find_id_column(df.columns)
    if id_column is None:
        raise ValueError('Spalte "ID" nicht gefunden.')

    id_values = df[id_column]
    has_id = id_values.notna() & (id_values.astype(str).str.strip() != "")
    return df.loc[has_id]


def _flat_welding_param_columns():
    return [
        field_name
        for block in WELDING_PARAM_TABLE_COLUMNS
        for field_name in block
    ]


def _required_excel_column_names():
    required = set(
        MENU_TABLE_COLUMNS
        + MENU_FILTER_OPTIONS
        + WELDING_FORM_COLUMNS
        + WELDING_METADATA_COLUMNS
        + SCANNING_METADATA_COLUMNS
        + ANALYZING_METADATA_COLUMNS
        + _flat_welding_param_columns()
    )
    return sorted(required)


def _get_experiment_fields(row, field_names):
    if versuchsuebersicht_data is None or row is None:
        return {}

    columns = versuchsuebersicht_data["columns"]
    fields = {}
    for name in field_names:
        column = _find_column(columns, name)
        if not column:
            fields[name] = ""
            continue
        fields[name] = _format_display_value(name, row.get(column))
    return fields


def _parse_positive_int_setting(value, field_name, minimum=1):
    text = _format_cell_value(value)
    if not text:
        raise ValueError(f"{field_name} fehlt.")

    normalized = text.replace(",", ".")
    try:
        parsed = int(float(normalized))
    except ValueError as exc:
        raise ValueError(f"{field_name} muss eine ganze Zahl sein.") from exc

    if parsed < minimum:
        raise ValueError(f"{field_name} muss >= {minimum} sein.")

    return parsed


def _parse_positive_float_setting(value, field_name):
    text = _format_cell_value(value)
    if not text:
        raise ValueError(f"{field_name} fehlt.")

    normalized = text.replace(",", ".")
    try:
        parsed = float(normalized)
    except ValueError as exc:
        raise ValueError(f"{field_name} muss eine Zahl sein.") from exc

    if parsed <= 0:
        raise ValueError(f"{field_name} muss größer als 0 sein.")

    return parsed


def _resolve_profile_frequency_from_row(row, columns):
    freq_col = _find_column(columns, "PROFILEFREQUENCY")
    if not freq_col:
        raise ValueError("PROFILEFREQUENCY fehlt.")
    return _parse_positive_float_setting(row.get(freq_col), "PROFILEFREQUENCY")


def _scan_settings_dict_from_values(
    *,
    resolution,
    profile_frequency,
    exposure_us,
    scan_duration_s,
    scan_duration_auto=False,
):
    run_scan = _import_run_scan()
    settings = run_scan.ScanSettings.from_profile_frequency(
        resolution=resolution,
        profile_frequency=profile_frequency,
        exposure_us=exposure_us,
        scan_duration_s=scan_duration_s,
        scan_duration_auto=scan_duration_auto,
    )
    return {
        "resolution": settings.resolution,
        "profile_frequency": settings.profile_frequency,
        "target_profile_count": settings.target_profile_count,
        "exposure_us": settings.exposure_us,
        "idle_time_us": settings.idle_time_us,
        "scan_duration_s": settings.scan_duration_s,
        "scan_duration_auto": settings.scan_duration_auto,
    }


def _get_scan_settings_from_row(row):
    if versuchsuebersicht_data is None or row is None:
        raise ValueError("Keine Versuchsübersicht geladen.")

    columns = versuchsuebersicht_data["columns"]
    resolution = _parse_positive_int_setting(
        row.get(_find_column(columns, "RESOLUTION")),
        "RESOLUTION",
    )
    scan_duration_s = _parse_positive_float_setting(
        row.get(_find_column(columns, "SCANDURATION")),
        "SCANDURATION",
    )
    profile_frequency = _resolve_profile_frequency_from_row(row, columns)

    return _scan_settings_dict_from_values(
        resolution=resolution,
        profile_frequency=profile_frequency,
        exposure_us=SCAN_DEFAULT_EXPOSURE_US,
        scan_duration_s=scan_duration_s,
    )


def _default_scan_settings():
    return _scan_settings_dict_from_values(
        resolution=640,
        profile_frequency=10.0,
        exposure_us=SCAN_DEFAULT_EXPOSURE_US,
        scan_duration_s=1.0,
    )


def _parse_bool_setting(value):
    if isinstance(value, bool):
        return value
    text = _format_cell_value(value).strip().lower()
    return text in {"1", "true", "yes", "on"}


def _scan_settings_to_excel_fields(scan_settings, auto_scan_duration_s=None, scanspeed=None):
    fields = {
        "RESOLUTION": scan_settings["resolution"],
        "PROFILEFREQUENCY": scan_settings["profile_frequency"],
    }

    if scan_settings.get("scan_duration_auto"):
        if auto_scan_duration_s is not None:
            fields["SCANDURATION"] = auto_scan_duration_s
    else:
        fields["SCANDURATION"] = scan_settings["scan_duration_s"]

    if scanspeed is not None:
        fields[SCANSPEED_COLUMN] = scanspeed

    return fields


def _get_scanspeed_from_row(row):
    if versuchsuebersicht_data is None or row is None:
        return ""

    columns = versuchsuebersicht_data["columns"]
    column = _find_column(columns, SCANSPEED_COLUMN)
    if not column:
        return ""

    return _format_cell_value(row.get(column))


def _resolve_scanspeed_value(scanspeed, row=None):
    if scanspeed is not None:
        text = str(scanspeed).strip()
        if not text:
            raise ValueError(f"{SCANSPEED_COLUMN} fehlt.")
        return _parse_positive_float_setting(text, SCANSPEED_COLUMN)

    value = _get_scanspeed_from_row(row)
    if not value:
        raise ValueError(f"{SCANSPEED_COLUMN} fehlt.")
    return _parse_positive_float_setting(value, SCANSPEED_COLUMN)


def _get_scan_geometry_from_row(row):
    if versuchsuebersicht_data is None or row is None:
        raise ValueError("Keine Versuchsübersicht geladen.")

    columns = versuchsuebersicht_data["columns"]
    scan_duration_s = _parse_positive_float_setting(
        row.get(_find_column(columns, "SCANDURATION")),
        "SCANDURATION",
    )
    scan_speed_mm_s = _resolve_scanspeed_value(_get_scanspeed_from_row(row), row)
    return scan_speed_mm_s, scan_duration_s


def _get_profile_y_mm_values(document, profiles, experiment_id=None):
    run_scan = _import_run_scan()

    if profiles and all(profile.get("y_mm") is not None for profile in profiles):
        return [float(profile["y_mm"]) for profile in profiles]

    scan_speed_mm_s = document.get("scan_speed_mm_s")
    scan_duration_s = document.get("scan_duration_s")
    scan_settings = document.get("scan_settings") or {}
    if scan_speed_mm_s is None:
        scan_speed_mm_s = scan_settings.get("scan_speed_mm_s")
    if scan_duration_s is None:
        try:
            scan_duration_s = run_scan.extract_scan_duration_s(scan_settings)
        except ValueError:
            scan_duration_s = None

    normalized_id = _normalize_id(experiment_id or document.get("experiment_id"))
    if normalized_id and versuchsuebersicht_data:
        row = versuchsuebersicht_data["id_index"].get(normalized_id)
        if row is not None:
            try:
                excel_speed, excel_duration = _get_scan_geometry_from_row(row)
                if scan_speed_mm_s is None:
                    scan_speed_mm_s = excel_speed
                if scan_duration_s is None:
                    scan_duration_s = excel_duration
            except ValueError:
                pass

    if scan_speed_mm_s is None or scan_duration_s is None:
        raise ValueError("SCANSPEED oder SCANDURATION fehlen für y-Zuordnung.")

    temp_profiles = [{"profile_index": index} for index in range(len(profiles))]
    run_scan.assign_profile_y_mm(
        temp_profiles,
        scan_speed_mm_s=float(scan_speed_mm_s),
        scan_duration_s=float(scan_duration_s),
    )
    return [float(item["y_mm"]) for item in temp_profiles]


def _get_scan_settings_from_ui(payload):
    if not isinstance(payload, dict):
        raise ValueError("Ungültige Scan-Einstellungen.")

    scan_duration_auto = _parse_bool_setting(payload.get("scan_duration_auto"))
    if scan_duration_auto:
        scan_duration_s = SCAN_DURATION_AUTO_SECONDS
    else:
        scan_duration_s = _parse_positive_float_setting(
            payload.get("scan_duration_s"),
            "SCANDURATION",
        )

    return _scan_settings_dict_from_values(
        resolution=_parse_positive_int_setting(payload.get("resolution"), "RESOLUTION"),
        profile_frequency=_parse_positive_float_setting(
            payload.get("profile_frequency"),
            "PROFILEFREQUENCY",
        ),
        exposure_us=SCAN_DEFAULT_EXPOSURE_US,
        scan_duration_s=scan_duration_s,
        scan_duration_auto=scan_duration_auto,
    )


def _get_scan_settings_for_experiment(experiment_id):
    if versuchsuebersicht_data is None:
        raise ValueError("Keine Versuchsübersicht geladen.")

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        raise ValueError("ID nicht gefunden.")

    try:
        return _get_scan_settings_from_row(row), False
    except ValueError:
        return _default_scan_settings(), True


def _get_live_profile_settings(experiment_id):
    query = request.args
    required = ("resolution", "profile_frequency", "scan_duration_s")
    if all(query.get(name) not in (None, "") for name in required):
        try:
            resolution = _parse_positive_int_setting(query.get("resolution"), "RESOLUTION")
            profile_frequency = _parse_positive_float_setting(
                query.get("profile_frequency"),
                "PROFILEFREQUENCY",
            )
            scan_duration_s = _parse_positive_float_setting(
                query.get("scan_duration_s"),
                "SCANDURATION",
            )
            return (
                _scan_settings_dict_from_values(
                    resolution=resolution,
                    profile_frequency=profile_frequency,
                    exposure_us=SCAN_DEFAULT_EXPOSURE_US,
                    scan_duration_s=scan_duration_s,
                ),
                False,
            )
        except ValueError:
            pass

    return _get_scan_settings_for_experiment(experiment_id)


def _import_run_scan():
    scripts_dir = str(BASE_DIR / "scripts")
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)
    import run_scan

    return run_scan


def _import_run_weld():
    scripts_dir = str(BASE_DIR / "scripts")
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)
    import run_weld

    return run_weld


def _import_run_analyze():
    scripts_dir = str(BASE_DIR / "scripts")
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)
    import run_analyze

    return run_analyze


def _close_live_profile_preview():
    _import_run_scan().close_live_profile_preview()


def _scan_settings_object(settings_dict):
    run_scan = _import_run_scan()
    return run_scan.ScanSettings.from_values(
        resolution=int(settings_dict["resolution"]),
        profile_frequency=float(settings_dict["profile_frequency"]),
        exposure_us=int(settings_dict["exposure_us"]),
        idle_time_us=int(settings_dict["idle_time_us"]),
        scan_duration_s=float(settings_dict["scan_duration_s"]),
        scan_duration_auto=bool(settings_dict.get("scan_duration_auto")),
    )


def _build_profile_graph_payload(
    profile,
    *,
    live=False,
    demo_mode=False,
    uses_default_settings=False,
    scanning=False,
):
    point_stats = _profile_point_stats(profile)
    x_mm, z_mm = _filter_valid_profile_points(
        profile["x_mm"],
        profile["z_mm"],
        profile.get("intensities"),
    )
    x_mm, z_mm = _downsample_profile_points(
        x_mm,
        z_mm,
        SCANNING_GRAPH_MAX_POINTS,
    )

    payload = {
        "profile_index": 0,
        "profile_count": 1,
        "resolution": int(profile.get("resolution") or len(profile.get("x_mm", []))),
        "demo_mode": demo_mode,
        "point_stats": point_stats,
        "x_mm": x_mm.tolist(),
        "z_mm": z_mm.tolist(),
    }
    if live:
        payload["live"] = True
        payload["uses_default_settings"] = uses_default_settings
    if scanning:
        payload["scanning"] = True
    return payload


def _build_weld_config():
    run_weld = _import_run_weld()
    return run_weld.WeldConfig(
        device=WELD_DEVICE,
        rate=WELD_RATE,
        threshold_v=WELD_THRESHOLD_V,
        pre_time_s=WELD_PRE_TIME_S,
        post_time_s=WELD_POST_TIME_S,
        max_duration_s=WELD_MAX_DURATION_S,
    )


def _get_weld_job(experiment_id):
    normalized_id = _normalize_id(experiment_id)
    with _weld_jobs_lock:
        return _weld_jobs.get(normalized_id)


def _set_weld_job(experiment_id, payload):
    normalized_id = _normalize_id(experiment_id)
    with _weld_jobs_lock:
        _weld_jobs[normalized_id] = payload


def _clear_weld_job(experiment_id):
    normalized_id = _normalize_id(experiment_id)
    with _weld_jobs_lock:
        _weld_jobs.pop(normalized_id, None)
        _weld_abort_events.pop(normalized_id, None)


def _run_weld_job(experiment_id, use_demo=False):
    normalized_id = _normalize_id(experiment_id)
    abort_event = threading.Event()
    with _weld_jobs_lock:
        _weld_abort_events[normalized_id] = abort_event

    try:
        _set_weld_job(normalized_id, {"status": "running"})
        run_weld = _import_run_weld()
        result = run_weld.run_weld(
            normalized_id,
            WELDING_DATA_DIR,
            demo=use_demo,
            config=_build_weld_config(),
            abort_event=abort_event,
        )
        updated_fields = _mark_experiment_welded_in_excel(
            Path(versuchsuebersicht_path),
            normalized_id,
            result["h5_file"],
        )
        _load_versuchsuebersicht_from_path(Path(versuchsuebersicht_path))
        result.pop("success", None)
        _set_weld_job(
            normalized_id,
            {
                "status": "done",
                "result": result,
                "fields": updated_fields,
            },
        )
    except Exception as exc:
        _set_weld_job(normalized_id, {"status": "error", "error": str(exc)})
    finally:
        with _weld_jobs_lock:
            _weld_abort_events.pop(normalized_id, None)


def _get_scan_job(experiment_id):
    normalized_id = _normalize_id(experiment_id)
    with _scan_jobs_lock:
        return _scan_jobs.get(normalized_id)


def _set_scan_job(experiment_id, payload):
    normalized_id = _normalize_id(experiment_id)
    with _scan_jobs_lock:
        _scan_jobs[normalized_id] = payload


def _clear_scan_job(experiment_id):
    normalized_id = _normalize_id(experiment_id)
    with _scan_jobs_lock:
        _scan_jobs.pop(normalized_id, None)


def _capture_live_profile_for_experiment(experiment_id):
    run_scan = _import_run_scan()
    normalized_id = _normalize_id(experiment_id)
    job = _get_scan_job(experiment_id)
    if job and job.get("status") == "running":
        cached = run_scan.get_scan_progress_profile(normalized_id)
        if cached is None:
            raise ValueError("Scan läuft – warte auf erstes Profil...")
        return cached, run_scan.get_scan_progress_demo_mode(normalized_id), False

    settings_dict, uses_default_settings = _get_live_profile_settings(experiment_id)
    settings = _scan_settings_object(settings_dict)

    if SCAN_USE_DEMO:
        profile = run_scan.capture_demo_live_profile(settings)
        return profile, True, uses_default_settings

    profile = run_scan.capture_live_profile(SCANNER_IP, settings)
    return profile, False, uses_default_settings


def _extract_auto_scan_duration_s(scan_result):
    scan_settings = scan_result.get("scan_settings") or {}
    if not scan_settings.get("scan_duration_auto"):
        return None

    scan_window = scan_settings.get("scan_window") or {}
    duration_s = scan_window.get("effective_duration_s", scan_settings.get("scan_duration_s"))
    if duration_s is None:
        return None

    return round(float(duration_s), 1)


def _run_scan_job(experiment_id, scan_settings, use_demo=False, scanspeed=None):
    normalized_id = _normalize_id(experiment_id)
    try:
        _set_scan_job(normalized_id, {"status": "running"})
        row = versuchsuebersicht_data["id_index"].get(normalized_id) if versuchsuebersicht_data else None
        resolved_scanspeed = _resolve_scanspeed_value(scanspeed, row)
        if versuchsuebersicht_path is not None:
            excel_path = Path(versuchsuebersicht_path)
            _repair_orphaned_scan_metadata(normalized_id)
            _save_scanning_fields_to_excel(
                excel_path,
                normalized_id,
                _scan_settings_to_excel_fields(scan_settings, scanspeed=resolved_scanspeed),
                editable_columns=SCANNING_PRESCAN_EXCEL_COLUMNS,
            )
            _load_versuchsuebersicht_from_path(excel_path)

        run_scan = _import_run_scan()
        run_scan.close_live_profile_preview()
        settings = _scan_settings_object(scan_settings)
        result = run_scan.run_scan(
            normalized_id,
            SCANNING_DATA_DIR,
            demo=use_demo,
            scanner_ip=SCANNER_IP,
            settings=settings,
            scan_speed_mm_s=resolved_scanspeed,
        )
        auto_scan_duration_s = _extract_auto_scan_duration_s(result)
        updated_fields = _mark_experiment_scanned_in_excel(
            Path(versuchsuebersicht_path),
            normalized_id,
            result["json_file"],
            scan_settings=scan_settings,
            auto_scan_duration_s=auto_scan_duration_s,
            scanspeed=resolved_scanspeed,
        )
        _load_versuchsuebersicht_from_path(Path(versuchsuebersicht_path))
        result.pop("success", None)
        _set_scan_job(
            normalized_id,
            {
                "status": "done",
                "result": result,
                "fields": updated_fields,
                "scan_settings": scan_settings,
            },
        )
    except Exception as exc:
        _set_scan_job(normalized_id, {"status": "error", "error": str(exc)})


def _is_default_excel_placeholder_column(name):
    return _DEFAULT_EXCEL_PLACEHOLDER_COLUMN.match(str(name).strip()) is not None


def _get_excel_usecols(all_columns):
    selected = set()

    for index, column in enumerate(all_columns):
        if _is_default_excel_placeholder_column(column):
            break
        selected.add(index)

    for name in _required_excel_column_names():
        for index, column in enumerate(all_columns):
            if str(column).strip().lower() == name.strip().lower():
                selected.add(index)

    return sorted(selected)


def _load_versuchsuebersicht_from_path(path):
    global versuchsuebersicht_path, versuchsuebersicht_data

    if path.suffix.lower() not in EXCEL_SUFFIXES:
        raise ValueError("Nur Excel-Dateien (.xlsx, .xls, .xlsm) werden unterstützt.")

    path = Path(path)

    header = pd.read_excel(path, nrows=0)
    all_columns = [str(column) for column in header.columns]
    df = pd.read_excel(path, nrows=MAX_ROWS, usecols=_get_excel_usecols(all_columns))
    df.columns = [str(column) for column in df.columns]
    df = _filter_rows_with_id(df)
    records = df.where(pd.notnull(df), None).to_dict(orient="records")
    columns = list(df.columns)

    versuchsuebersicht_path = str(path.resolve())
    versuchsuebersicht_data = {
        "columns": columns,
        "rows": records,
        "id_index": _build_id_index(records, columns),
    }

    return {
        "path": versuchsuebersicht_path,
        "row_count": len(records),
        "column_count": len(columns),
    }


def _try_load_default_versuchsuebersicht():
    if not DEFAULT_VERSUCHSUEBERSICHT.is_file():
        return

    try:
        _load_versuchsuebersicht_from_path(DEFAULT_VERSUCHSUEBERSICHT)
    except Exception:
        pass


def _get_load_status():
    if versuchsuebersicht_data is None:
        return {"message": "Noch nicht geladen.", "type": "idle"}

    row_count = len(versuchsuebersicht_data["rows"])
    return {"message": f"{row_count} Zeilen geladen.", "type": "success"}


def _get_versuchs_path_display():
    if versuchsuebersicht_path:
        path = Path(versuchsuebersicht_path)
        try:
            return str(path.relative_to(BASE_DIR))
        except ValueError:
            return versuchsuebersicht_path

    return DEFAULT_VERSUCHSUEBERSICHT_PATH


def _has_value(value):
    if value is None:
        return False

    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass

    text = str(value).strip()
    if not text:
        return False

    return text.lower() not in {"nan", "nat", "none"}


def _format_cell_value(value):
    if not _has_value(value):
        return ""

    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]

    return text


def _format_welded(value):
    if not _has_value(value):
        return ""

    parsed = _parse_excel_timestamp(value)
    if parsed is not None:
        return parsed.strftime("%Y-%m-%d %H:%M:%S")

    text = _format_cell_value(value)
    return text


def _format_display_value(column_name, value):
    if column_name in {"WELDED", "SCANNED"}:
        return _format_welded(value)
    return _format_cell_value(value)


def _get_filter_options(columns):
    options = []
    for name in MENU_FILTER_OPTIONS:
        column = _find_column(columns, name)
        if column:
            options.append(column)
    return options


def _resolve_filter_column(requested_column):
    if versuchsuebersicht_data is None:
        return DEFAULT_TABLE_FILTER_COLUMN

    columns = versuchsuebersicht_data["columns"]
    options = _get_filter_options(columns)

    if not options:
        return DEFAULT_TABLE_FILTER_COLUMN

    if requested_column:
        matched_column = _find_column(columns, requested_column)
        if matched_column in options:
            return matched_column

    matched_default = _find_column(columns, DEFAULT_TABLE_FILTER_COLUMN)
    return matched_default if matched_default in options else options[0]


def _get_table_rows(filter_column_name):
    if versuchsuebersicht_data is None:
        return []

    columns = versuchsuebersicht_data["columns"]
    column_map = {name: _find_column(columns, name) for name in MENU_TABLE_COLUMNS}
    filter_column = _find_column(columns, filter_column_name)

    if filter_column is None:
        return []

    rows = []
    for row in versuchsuebersicht_data["rows"]:
        if not _has_value(row.get(filter_column)):
            continue

        display_row = {}
        for name in MENU_TABLE_COLUMNS:
            source_column = column_map[name]
            if not source_column:
                display_row[name] = ""
                continue

            display_row[name] = _format_display_value(name, row.get(source_column))

        rows.append(display_row)

    return rows


def _get_analyzing_table_rows():
    if versuchsuebersicht_data is None:
        return []

    columns = versuchsuebersicht_data["columns"]
    column_map = {name: _find_column(columns, name) for name in ANALYZING_TABLE_COLUMNS}
    welded_column = _find_column(columns, "WELDED")
    scanned_column = _find_column(columns, "SCANNED")

    if welded_column is None or scanned_column is None:
        return []

    rows = []
    for row in versuchsuebersicht_data["rows"]:
        if not _has_value(row.get(welded_column)):
            continue
        if not _has_value(row.get(scanned_column)):
            continue

        display_row = {}
        for name in ANALYZING_TABLE_COLUMNS:
            source_column = column_map[name]
            if not source_column:
                display_row[name] = ""
                continue

            display_row[name] = _format_display_value(name, row.get(source_column))
            if name in {"WELDED", "SCANNED"}:
                parsed = _parse_excel_timestamp(row.get(source_column))
                if parsed is not None:
                    display_row[name] = parsed.strftime("%d.%m.%y %H:%M")

        experiment_id = display_row.get("ID", "")
        display_row["has_analyze"] = (
            _experiment_has_valid_analyze(experiment_id) if experiment_id else False
        )

        rows.append(display_row)

    return rows


def _build_excel_header_map(worksheet):
    header_map = {}
    for col_idx in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=col_idx).value
        if value is not None and str(value).strip():
            header_map[str(value).strip().lower()] = col_idx
    return header_map


def _parse_excel_timestamp(value):
    if not _has_value(value):
        return None

    if isinstance(value, datetime):
        return value.replace(tzinfo=None)

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().replace(tzinfo=None)

    text = str(value).strip()
    if not text:
        return None

    try:
        return pd.Timestamp(text).to_pydatetime().replace(tzinfo=None)
    except (ValueError, TypeError):
        return None


def _find_excel_row_by_id(worksheet, header_map, experiment_id):
    id_col = header_map.get("id")
    if id_col is None:
        raise ValueError('Spalte "ID" nicht gefunden.')

    normalized_id = _normalize_id(experiment_id)
    for row_idx in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_idx, column=id_col).value
        if _normalize_id(value) == normalized_id:
            return row_idx

    return None


def _excel_cell_value(field_name, value):
    if value == "" or value is None:
        return None

    if field_name not in EXCEL_NUMERIC_COLUMNS:
        return value

    if isinstance(value, bool):
        return value

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if field_name in EXCEL_INTEGER_COLUMNS and value.is_integer():
            return int(value)
        return value

    normalized = str(value).strip().replace(",", ".")
    if not normalized:
        return None

    try:
        number = float(normalized)
    except ValueError:
        return value

    if field_name in EXCEL_INTEGER_COLUMNS and number.is_integer():
        return int(number)
    return number


def _validate_welding_field_values(field_values):
    path_id = str(field_values.get("PATH ID", "")).strip().upper()
    if path_id and path_id not in PATH_ID_OPTIONS:
        raise ValueError("PATH ID muss SO, PO, SI, ME oder LI sein.")


def _sanitize_h5_filename(filename):
    text = str(filename).strip()
    if not text:
        return None

    name = Path(text).name
    if not re.fullmatch(r"[\w.-]+\.h5", name, re.IGNORECASE):
        raise ValueError(f"Ungültiger H5-Dateiname: {name}")

    return name


def _get_experiment_h5_filename(experiment_id):
    if versuchsuebersicht_data is None:
        return None

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return None

    column = _find_column(versuchsuebersicht_data["columns"], "H5FILE")
    if column is None:
        return None

    value = _format_cell_value(row.get(column))
    if not value:
        return None

    return _sanitize_h5_filename(value)


def _sanitize_scan_filename(filename):
    text = str(filename).strip()
    if not text:
        return None

    name = Path(text).name
    if not re.fullmatch(r"[\w.-]+\.json", name, re.IGNORECASE):
        raise ValueError(f"Ungültiger Scan-Dateiname: {name}")

    return name


def _get_experiment_scan_filename(experiment_id):
    if versuchsuebersicht_data is None:
        return None

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return None

    column = _find_column(versuchsuebersicht_data["columns"], "JSONFILE")
    if column is None:
        return None

    value = _format_cell_value(row.get(column))
    if not value:
        return None

    return _sanitize_scan_filename(value)


def _load_scan_document(scan_path):
    cache_key = str(scan_path.resolve())
    mtime = scan_path.stat().st_mtime
    cached = _scan_document_cache.get(cache_key)
    if cached and cached[0] == mtime:
        return cached[1]

    with scan_path.open("r", encoding="utf-8") as handle:
        document = json.load(handle)

    if "profiles" not in document or not document["profiles"]:
        raise ValueError("Keine Profile in Scan-Datei gefunden.")

    _scan_document_cache[cache_key] = (mtime, document)
    return document


def _profile_valid_point_mask(profile):
    x_mm = np.asarray(profile["x_mm"], dtype=float)
    z_mm = np.asarray(profile["z_mm"], dtype=float)
    intensities = profile.get("intensities")

    if intensities is not None:
        return np.asarray(intensities, dtype=float) > 0

    return (np.abs(x_mm) > 1e-6) | (np.abs(z_mm) > 1e-6)


def _profile_point_stats(profile):
    valid_mask = _profile_valid_point_mask(profile)
    total_points = int(valid_mask.size)
    valid_points = int(valid_mask.sum())
    return {
        "total_points": total_points,
        "valid_points": valid_points,
        "invalid_points": total_points - valid_points,
    }


def _filter_valid_profile_points(x_mm, z_mm, intensities=None):
    x_mm = np.asarray(x_mm, dtype=float)
    z_mm = np.asarray(z_mm, dtype=float)

    if intensities is not None:
        intensities = np.asarray(intensities, dtype=float)
        mask = intensities > 0
    else:
        mask = (np.abs(x_mm) > 1e-6) | (np.abs(z_mm) > 1e-6)

    if not np.any(mask):
        return x_mm[:0], z_mm[:0]

    return x_mm[mask], z_mm[mask]


def _downsample_profile_points(x_mm, z_mm, max_points):
    x_mm = np.asarray(x_mm, dtype=float)
    z_mm = np.asarray(z_mm, dtype=float)

    if len(x_mm) <= max_points:
        return x_mm, z_mm

    indices = np.linspace(0, len(x_mm) - 1, max_points, dtype=int)
    return x_mm[indices], z_mm[indices]


def _load_scanning_profile_graph(scan_path, profile_index):
    document = _load_scan_document(scan_path)
    profiles = document["profiles"]
    profile_count = int(document.get("profile_count") or len(profiles))

    if profile_index < 0 or profile_index >= len(profiles):
        raise ValueError(f"Profilindex {profile_index} liegt außerhalb des Bereichs 0–{len(profiles) - 1}.")

    profile = profiles[profile_index]
    point_stats = _profile_point_stats(profile)
    y_mm = profile.get("y_mm")
    if y_mm is None:
        try:
            y_mm = _get_profile_y_mm_values(
                document,
                profiles,
                experiment_id=document.get("experiment_id"),
            )[profile_index]
        except ValueError:
            y_mm = None
    x_mm, z_mm = _filter_valid_profile_points(
        profile["x_mm"],
        profile["z_mm"],
        profile.get("intensities"),
    )
    x_mm, z_mm = _downsample_profile_points(
        x_mm,
        z_mm,
        SCANNING_GRAPH_MAX_POINTS,
    )

    return {
        "json_file": scan_path.name,
        "profile_index": profile_index,
        "profile_count": profile_count,
        "resolution": int(document.get("resolution") or len(profile.get("x_mm", []))),
        "demo_mode": bool(document.get("demo_mode")),
        "point_stats": point_stats,
        "y_mm": y_mm,
        "x_mm": x_mm.tolist(),
        "z_mm": z_mm.tolist(),
    }


def _load_analyzing_scan_heatmap(analyze_path):
    document = _load_scan_document(analyze_path)
    profiles = document.get("profiles") or []
    if not profiles:
        raise ValueError("Keine Profile in Analyse-Datei gefunden.")

    profile_count = len(profiles)
    resolution = int(document.get("resolution") or len(profiles[0].get("x_mm", [])))
    x_mm = profiles[0].get("x_mm") or []
    if not x_mm:
        raise ValueError("Keine x-Werte in Analyse-Datei gefunden.")

    x_first = float(x_mm[0])
    x_last = float(x_mm[-1])
    x_min = min(x_first, x_last)
    x_max = max(x_first, x_last)

    values = []
    positive_values = []
    for profile in profiles:
        z_row = [float(value) for value in profile.get("z_mm") or []]
        if len(z_row) != resolution:
            raise ValueError("Profilauflösung ist in der Analyse-Datei nicht konsistent.")
        values.extend(z_row)
        positive_values.extend(value for value in z_row if value > 0)

    z_scale_min = 0.0
    z_scale_max = 0.0
    if positive_values:
        sorted_values = sorted(positive_values)
        lower_index = max(0, int(len(sorted_values) * 0.02) - 1)
        upper_index = min(len(sorted_values) - 1, int(len(sorted_values) * 0.98))
        z_scale_min = float(sorted_values[lower_index])
        z_scale_max = float(sorted_values[upper_index])
        if z_scale_max <= z_scale_min:
            z_scale_min = float(sorted_values[0])
            z_scale_max = float(sorted_values[-1])

    y_mm = _get_profile_y_mm_values(
        document,
        profiles,
        experiment_id=document.get("experiment_id"),
    )

    return {
        "json_file": analyze_path.name,
        "profile_count": profile_count,
        "resolution": resolution,
        "x_min": x_min,
        "x_max": x_max,
        "x_reversed": x_first > x_last,
        "y_mm": y_mm,
        "y_min": min(y_mm),
        "y_max": max(y_mm),
        "z_scale_min": z_scale_min,
        "z_scale_max": z_scale_max,
        "values": values,
    }


def _resolve_scan_path_for_experiment(experiment_id):
    try:
        scan_filename = _get_experiment_scan_filename(experiment_id)
    except ValueError as exc:
        return None, str(exc)

    if not scan_filename:
        return None, "Keine Scan-Datei für diese ID."

    scan_path = SCANNING_DATA_DIR / scan_filename
    if not scan_path.is_file():
        return None, f"Scan-Datei nicht gefunden: {scan_filename}"

    return scan_path, None


def _sanitize_analyze_filename(filename):
    text = str(filename).strip()
    if not text:
        return None

    name = Path(text).name
    if not re.fullmatch(r"[\w.-]+\.json", name, re.IGNORECASE):
        raise ValueError(f"Ungültiger Analyse-Dateiname: {name}")

    return name


def _get_experiment_analyze_filename(experiment_id):
    if versuchsuebersicht_data is None:
        return None

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return None

    column = _find_column(versuchsuebersicht_data["columns"], "ANALYZESCAN")
    if column is None:
        return None

    value = _format_cell_value(row.get(column))
    if not value:
        return None

    return _sanitize_analyze_filename(value)


def _resolve_analyze_path_for_experiment(experiment_id):
    try:
        analyze_filename = _get_experiment_analyze_filename(experiment_id)
    except ValueError as exc:
        return None, str(exc)

    normalized_id = _normalize_id(experiment_id)
    candidates = []
    if analyze_filename:
        candidates.append(analyze_filename)
    default_filename = f"{normalized_id}.json"
    if default_filename not in candidates:
        candidates.append(default_filename)

    for filename in candidates:
        analyze_path = ANALYZING_DATA_DIR / filename
        if analyze_path.is_file():
            return analyze_path, None

    if analyze_filename:
        return None, f"Analyse-Datei nicht gefunden: {analyze_filename}"

    return None, "Keine Analyse-Datei für diese ID."


def _get_experiment_analyze_weld_filename(experiment_id):
    if versuchsuebersicht_data is None:
        return None

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return None

    column = _find_column(versuchsuebersicht_data["columns"], "ANALYZEWELD")
    if column is None:
        return None

    value = _format_cell_value(row.get(column))
    if not value:
        return None

    return _sanitize_h5_filename(value)


def _resolve_analyze_weld_path_for_experiment(experiment_id):
    try:
        analyze_weld_filename = _get_experiment_analyze_weld_filename(experiment_id)
    except ValueError as exc:
        return None, str(exc)

    normalized_id = _normalize_id(experiment_id)
    candidates = []
    if analyze_weld_filename:
        candidates.append(analyze_weld_filename)
    default_filename = f"{normalized_id}.h5"
    if default_filename not in candidates:
        candidates.append(default_filename)

    for filename in candidates:
        analyze_weld_path = ANALYZING_DATA_DIR / filename
        if analyze_weld_path.is_file():
            return analyze_weld_path, None

    if analyze_weld_filename:
        return None, f"Analyse-H5-Datei nicht gefunden: {analyze_weld_filename}"

    return None, "Keine Analyse-H5-Datei für diese ID."


def _experiment_has_valid_analyze(experiment_id):
    try:
        analyze_scan_filename = _get_experiment_analyze_filename(experiment_id)
        analyze_weld_filename = _get_experiment_analyze_weld_filename(experiment_id)
    except ValueError:
        return False

    if not analyze_scan_filename or not analyze_weld_filename:
        return False

    scan_path = ANALYZING_DATA_DIR / analyze_scan_filename
    weld_path = ANALYZING_DATA_DIR / analyze_weld_filename
    return scan_path.is_file() and weld_path.is_file()


def _excel_row_has_scan_metadata(row):
    if versuchsuebersicht_data is None or row is None:
        return False

    columns = versuchsuebersicht_data["columns"]
    scanned_column = _find_column(columns, "SCANNED")
    jsonfile_column = _find_column(columns, "JSONFILE")
    return (
        (scanned_column and _has_value(row.get(scanned_column)))
        or (jsonfile_column and _has_value(row.get(jsonfile_column)))
    )


def _experiment_has_valid_scan(experiment_id):
    scan_path, _error = _resolve_scan_path_for_experiment(experiment_id)
    return scan_path is not None


def _repair_orphaned_scan_metadata(experiment_id):
    if versuchsuebersicht_path is None or versuchsuebersicht_data is None:
        return None

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None or _experiment_has_valid_scan(experiment_id):
        return None

    if not _excel_row_has_scan_metadata(row):
        return None

    return _reset_experiment_scan_in_excel(Path(versuchsuebersicht_path), experiment_id)


def _apply_scan_display_fields(fields, experiment_id):
    if _experiment_has_valid_scan(experiment_id):
        return fields

    sanitized = dict(fields)
    sanitized["SCANNED"] = ""
    sanitized["JSONFILE"] = ""
    return sanitized


def _downsample_welding_series(time_s, channels, max_points):
    if len(time_s) <= max_points:
        return time_s, channels

    indices = np.linspace(0, len(time_s) - 1, max_points, dtype=int)
    downsampled_time = time_s[indices]
    downsampled_channels = []
    for channel in channels:
        downsampled_channels.append(
            {
                **channel,
                "values": channel["values"][indices],
            }
        )

    return downsampled_time, downsampled_channels


def _format_h5_channel_name(dataset, fallback_name):
    label = dataset.attrs.get("label")
    units = dataset.attrs.get("units")
    if label and units:
        return f"{label} [{units}]"
    if label:
        return str(label)
    return fallback_name


def _load_welding_graph_from_h5(h5_path):
    with h5py.File(h5_path, "r") as h5_file:
        if "time_s" not in h5_file:
            raise ValueError('Datensatz "time_s" nicht gefunden.')

        time_s = np.array(h5_file["time_s"], dtype=float)
        channels = []
        for key in sorted(name for name in h5_file.keys() if name.startswith("channel_")):
            dataset = h5_file[key]
            channels.append(
                {
                    "key": key,
                    "name": _format_h5_channel_name(dataset, key),
                    "values": np.array(dataset, dtype=float),
                }
            )

    if not channels:
        raise ValueError("Keine Messkanäle in H5-Datei gefunden.")

    time_s, channels = _downsample_welding_series(time_s, channels, WELDING_GRAPH_MAX_POINTS)
    return {
        "time_s": time_s.tolist(),
        "channels": [
            {
                "key": channel["key"],
                "name": channel["name"],
                "values": channel["values"].tolist(),
            }
            for channel in channels
        ],
    }


def _save_welding_fields_to_excel(path, experiment_id, field_values):
    workbook = load_workbook(path)
    worksheet = workbook.active
    header_map = _build_excel_header_map(worksheet)
    row_idx = _find_excel_row_by_id(worksheet, header_map, experiment_id)

    if row_idx is None:
        raise ValueError("ID nicht gefunden.")

    updated_fields = {}
    for name in WELDING_FORM_COLUMNS:
        if name not in field_values:
            continue

        col_idx = header_map.get(name.strip().lower())
        if col_idx is None:
            continue

        value = field_values[name]
        updated_fields[name] = _write_excel_form_field(
            worksheet,
            row_idx,
            col_idx,
            name,
            value,
        )

    workbook.save(path)
    return updated_fields


def _save_scanning_fields_to_excel(path, experiment_id, field_values, editable_columns=None):
    if editable_columns is None:
        editable_columns = SCANNING_EDITABLE_COLUMNS

    with _excel_workbook_lock:
        workbook = load_workbook(path)
        worksheet = workbook.active
        header_map = _build_excel_header_map(worksheet)
        row_idx = _find_excel_row_by_id(worksheet, header_map, experiment_id)

        if row_idx is None:
            raise ValueError("ID nicht gefunden.")

        updated_fields = {}
        for name in editable_columns:
            if name not in field_values:
                continue

            col_idx = header_map.get(name.strip().lower())
            if col_idx is None:
                continue

            value = field_values[name]
            updated_fields[name] = _write_excel_form_field(
                worksheet,
                row_idx,
                col_idx,
                name,
                value,
            )

        workbook.save(path)
        return updated_fields


def _mark_experiment_scanned_in_excel(
    path,
    experiment_id,
    scan_filename,
    scan_settings=None,
    auto_scan_duration_s=None,
    scanspeed=None,
):
    with _excel_workbook_lock:
        workbook = load_workbook(path)
        worksheet = workbook.active
        header_map = _build_excel_header_map(worksheet)
        row_idx = _find_excel_row_by_id(worksheet, header_map, experiment_id)

        if row_idx is None:
            raise ValueError("ID nicht gefunden.")

        scanned_col = header_map.get("scanned")
        jsonfile_col = header_map.get("jsonfile")
        scanned_at = datetime.now()

        if scanned_col is not None:
            worksheet.cell(row=row_idx, column=scanned_col, value=scanned_at)

        if jsonfile_col is not None:
            worksheet.cell(row=row_idx, column=jsonfile_col, value=scan_filename)

        updated_fields = {
            "SCANNED": scanned_at.strftime("%Y-%m-%d %H:%M:%S"),
            "JSONFILE": scan_filename,
        }

        if scan_settings is not None:
            settings_fields = _scan_settings_to_excel_fields(
                scan_settings,
                auto_scan_duration_s=auto_scan_duration_s,
                scanspeed=scanspeed,
            )
            for name, value in settings_fields.items():
                col_idx = header_map.get(name.strip().lower())
                if col_idx is None:
                    continue
                updated_fields[name] = _write_excel_form_field(
                    worksheet,
                    row_idx,
                    col_idx,
                    name,
                    value,
                )
        elif auto_scan_duration_s is not None:
            scanduration_col = header_map.get("scanduration")
            if scanduration_col is not None:
                updated_fields["SCANDURATION"] = _write_excel_form_field(
                    worksheet,
                    row_idx,
                    scanduration_col,
                    "SCANDURATION",
                    auto_scan_duration_s,
                )

        workbook.save(path)
        return updated_fields


def _mark_experiment_welded_in_excel(path, experiment_id, h5_filename):
    workbook = load_workbook(path)
    worksheet = workbook.active
    header_map = _build_excel_header_map(worksheet)
    row_idx = _find_excel_row_by_id(worksheet, header_map, experiment_id)

    if row_idx is None:
        raise ValueError("ID nicht gefunden.")

    welded_col = header_map.get("welded")
    h5file_col = header_map.get("h5file")
    welded_at = datetime.now()

    if welded_col is not None:
        worksheet.cell(row=row_idx, column=welded_col, value=welded_at)

    if h5file_col is not None:
        worksheet.cell(row=row_idx, column=h5file_col, value=h5_filename)

    workbook.save(path)
    return {
        "WELDED": welded_at.strftime("%Y-%m-%d %H:%M:%S"),
        "H5FILE": h5_filename,
    }


def _mark_experiment_analyzed_in_excel(
    path,
    experiment_id,
    analyze_scan_filename,
    analyze_weld_filename,
):
    workbook = load_workbook(path)
    worksheet = workbook.active
    header_map = _build_excel_header_map(worksheet)
    row_idx = _find_excel_row_by_id(worksheet, header_map, experiment_id)

    if row_idx is None:
        raise ValueError("ID nicht gefunden.")

    analyzescan_col = header_map.get("analyzescan")
    analyzeweld_col = header_map.get("analyzeweld")

    if analyzescan_col is not None and analyze_scan_filename is not None:
        worksheet.cell(row=row_idx, column=analyzescan_col, value=analyze_scan_filename)

    if analyzeweld_col is not None and analyze_weld_filename is not None:
        worksheet.cell(row=row_idx, column=analyzeweld_col, value=analyze_weld_filename)

    workbook.save(path)
    updated_fields = {}
    if analyze_scan_filename is not None:
        updated_fields["ANALYZESCAN"] = analyze_scan_filename
    if analyze_weld_filename is not None:
        updated_fields["ANALYZEWELD"] = analyze_weld_filename
    return updated_fields


def _is_blank_excel_input(value):
    if value is None:
        return True

    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass

    if isinstance(value, str) and not value.strip():
        return True

    return False


def _write_excel_form_field(worksheet, row_idx, col_idx, field_name, value):
    if field_name == "PATH ID":
        value = str(value).strip().upper()

    if _is_blank_excel_input(value):
        _clear_excel_cell(worksheet, row_idx, col_idx)
        return ""

    cell_value = _excel_cell_value(field_name, value)
    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
    if cell_value is None:
        return ""
    if isinstance(cell_value, datetime):
        return cell_value.strftime("%Y-%m-%d %H:%M:%S")
    return _format_cell_value(cell_value)


def _clear_excel_cell(worksheet, row_idx, col_idx):
    if col_idx is None:
        return
    worksheet.cell(row=row_idx, column=col_idx, value="")


def _reset_experiment_scan_in_excel(path, experiment_id):
    workbook = load_workbook(path)
    worksheet = workbook.active
    header_map = _build_excel_header_map(worksheet)
    row_idx = _find_excel_row_by_id(worksheet, header_map, experiment_id)

    if row_idx is None:
        raise ValueError("ID nicht gefunden.")

    scanned_col = header_map.get("scanned")
    jsonfile_col = header_map.get("jsonfile")
    deleted_file = None

    if jsonfile_col is not None:
        scan_filename = worksheet.cell(row=row_idx, column=jsonfile_col).value
        if _has_value(scan_filename):
            safe_filename = _sanitize_scan_filename(scan_filename)
            if safe_filename:
                scan_path = SCANNING_DATA_DIR / safe_filename
                if scan_path.is_file():
                    scan_path.unlink()
                    deleted_file = safe_filename
                    cache_key = str(scan_path.resolve())
                    _scan_document_cache.pop(cache_key, None)

        _clear_excel_cell(worksheet, row_idx, jsonfile_col)

    if scanned_col is not None:
        _clear_excel_cell(worksheet, row_idx, scanned_col)

    workbook.save(path)
    return {
        "SCANNED": "",
        "JSONFILE": "",
        "deleted_file": deleted_file,
    }


def _excel_row_has_weld_metadata(row):
    if versuchsuebersicht_data is None or row is None:
        return False

    columns = versuchsuebersicht_data["columns"]
    welded_column = _find_column(columns, "WELDED")
    h5file_column = _find_column(columns, "H5FILE")
    return (
        (welded_column and _has_value(row.get(welded_column)))
        or (h5file_column and _has_value(row.get(h5file_column)))
    )


def _resolve_h5_path_for_experiment(experiment_id):
    try:
        h5_filename = _get_experiment_h5_filename(experiment_id)
    except ValueError as exc:
        return None, str(exc)

    if not h5_filename:
        return None, "Keine H5-Datei für diese ID."

    h5_path = WELDING_DATA_DIR / h5_filename
    if not h5_path.is_file():
        return None, f"H5-Datei nicht gefunden: {h5_filename}"

    return h5_path, None


def _experiment_has_valid_weld(experiment_id):
    h5_path, _error = _resolve_h5_path_for_experiment(experiment_id)
    return h5_path is not None


def _reset_experiment_weld_in_excel(path, experiment_id):
    workbook = load_workbook(path)
    worksheet = workbook.active
    header_map = _build_excel_header_map(worksheet)
    row_idx = _find_excel_row_by_id(worksheet, header_map, experiment_id)

    if row_idx is None:
        raise ValueError("ID nicht gefunden.")

    welded_col = header_map.get("welded")
    h5file_col = header_map.get("h5file")
    deleted_file = None

    if h5file_col is not None:
        h5_filename = worksheet.cell(row=row_idx, column=h5file_col).value
        if _has_value(h5_filename):
            safe_filename = _sanitize_h5_filename(h5_filename)
            if safe_filename:
                h5_path = WELDING_DATA_DIR / safe_filename
                if h5_path.is_file():
                    h5_path.unlink()
                    deleted_file = safe_filename

        _clear_excel_cell(worksheet, row_idx, h5file_col)

    if welded_col is not None:
        _clear_excel_cell(worksheet, row_idx, welded_col)

    workbook.save(path)
    return {
        "WELDED": "",
        "H5FILE": "",
        "deleted_file": deleted_file,
    }


def _run_scan_script(experiment_id, use_demo=False, scan_settings=None):
    if not SCAN_SCRIPT_PATH.is_file():
        raise RuntimeError(f"Scan-Skript nicht gefunden: {SCAN_SCRIPT_PATH}")

    SCANNING_DATA_DIR.mkdir(parents=True, exist_ok=True)
    command = [
        sys.executable,
        str(SCAN_SCRIPT_PATH),
        "--experiment-id",
        experiment_id,
        "--output-dir",
        str(SCANNING_DATA_DIR),
    ]
    if use_demo:
        command.append("--demo")
    if SCANNER_IP:
        command.extend(["--scanner-ip", SCANNER_IP])
    if scan_settings:
        command.extend(
            [
                "--resolution",
                str(scan_settings["resolution"]),
                "--profile-frequency",
                str(scan_settings["profile_frequency"]),
                "--exposure-us",
                str(scan_settings["exposure_us"]),
                "--scan-duration-s",
                str(scan_settings["scan_duration_s"]),
            ]
        )
        if scan_settings.get("scan_duration_auto"):
            command.append("--scan-duration-auto")

    scan_timeout_s = 600 if scan_settings and scan_settings.get("scan_duration_auto") else 180
    completed = subprocess.run(
        command,
        capture_output=True,
        text=True,
        timeout=scan_timeout_s,
        cwd=str(BASE_DIR),
        check=False,
    )

    output_lines = [line.strip() for line in completed.stdout.splitlines() if line.strip()]
    if not output_lines:
        stderr = completed.stderr.strip()
        raise RuntimeError(stderr or "Scan-Skript ohne Ausgabe beendet.")

    payload = json.loads(output_lines[-1])
    if completed.returncode != 0 or not payload.get("success"):
        raise RuntimeError(payload.get("error") or completed.stderr.strip() or "Scan fehlgeschlagen.")

    return payload


@app.context_processor
def inject_nav_context():
    return {
        "load_status": _get_load_status(),
        "versuchs_path": _get_versuchs_path_display(),
    }


@app.get("/")
def index():
    return redirect(url_for("menu"))


@app.get("/menu")
def menu():
    filter_column = _resolve_filter_column(request.args.get("filter", DEFAULT_TABLE_FILTER_COLUMN))
    filter_options = (
        _get_filter_options(versuchsuebersicht_data["columns"])
        if versuchsuebersicht_data
        else []
    )
    table_rows = _get_table_rows(filter_column)

    return render_template(
        "menu.html",
        active_page="menu",
        table_rows=table_rows,
        table_columns=MENU_TABLE_COLUMNS,
        filter_column=filter_column,
        filter_options=filter_options,
    )


@app.get("/welding")
def welding():
    return render_template(
        "welding.html",
        active_page="welding",
        welding_param_table_columns=WELDING_PARAM_TABLE_COLUMNS,
        welding_editable_param_columns=[
            name
            for name in _flat_welding_param_columns()
            if name not in WELDING_PARAM_READONLY_COLUMNS
        ],
        resolution_options=SCAN_RESOLUTION_OPTIONS,
    )


@app.get("/scanning")
def scanning():
    return render_template(
        "scanning.html",
        active_page="scanning",
        scanning_param_table_columns=WELDING_PARAM_TABLE_COLUMNS,
        scanning_editable_param_columns=[
            name
            for name in _flat_welding_param_columns()
            if name not in SCANNING_PARAM_READONLY_COLUMNS
        ],
        scan_settings_required_columns=SCAN_SETTINGS_COLUMNS,
        scanning_editable_columns=SCANNING_EDITABLE_COLUMNS,
        scanning_editable_after_scan_columns=SCANNING_EDITABLE_AFTER_SCAN_COLUMNS,
        scanspeed_field=SCANSPEED_COLUMN,
        scan_default_exposure_us=SCAN_DEFAULT_EXPOSURE_US,
        resolution_options=SCAN_RESOLUTION_OPTIONS,
        scan_duration_auto_seconds=SCAN_DURATION_AUTO_SECONDS,
    )


@app.get("/analyzing")
def analyzing():
    table_rows = _get_analyzing_table_rows()

    return render_template(
        "analyzing.html",
        active_page="analyzing",
        table_rows=table_rows,
        table_columns=ANALYZING_TABLE_COLUMNS,
    )


@app.post("/api/analyzing/generate/<experiment_id>")
def generate_analyzing_data(experiment_id):
    if versuchsuebersicht_path is None or versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return jsonify(error="ID nicht gefunden."), 404

    scan_path = None
    scan_error = None
    if _experiment_has_valid_scan(experiment_id):
        scan_path, scan_error = _resolve_scan_path_for_experiment(experiment_id)

    h5_path = None
    h5_error = None
    if _experiment_has_valid_weld(experiment_id):
        h5_path, h5_error = _resolve_h5_path_for_experiment(experiment_id)

    if scan_path is None and h5_path is None:
        messages = []
        if scan_error:
            messages.append(scan_error)
        if h5_error:
            messages.append(h5_error)
        if not messages:
            messages.append("Für diese ID liegt keine gültige Scan- oder H5-Datei vor.")
        return jsonify(error=" ".join(messages)), 400

    try:
        scan_speed_mm_s = None
        scan_duration_s = None
        if scan_path is not None:
            scan_speed_mm_s, scan_duration_s = _get_scan_geometry_from_row(row)

        result = _import_run_analyze().run_analyze(
            experiment_id,
            ANALYZING_DATA_DIR,
            scan_path=scan_path,
            weld_path=h5_path,
            weld_config=_build_weld_config(),
            scan_speed_mm_s=scan_speed_mm_s,
            scan_duration_s=scan_duration_s,
        )
        updated_fields = _mark_experiment_analyzed_in_excel(
            Path(versuchsuebersicht_path),
            experiment_id,
            result.get("json_file"),
            result.get("h5_file"),
        )
        _load_versuchsuebersicht_from_path(Path(versuchsuebersicht_path))
        result["fields"] = updated_fields
        result["has_analyze"] = _experiment_has_valid_analyze(experiment_id)
        return jsonify(**result)
    except Exception as exc:
        return jsonify(error=f"Probendaten konnten nicht generiert werden: {exc}"), 500


@app.get("/api/analyzing/<experiment_id>/graph")
def get_analyzing_graph(experiment_id):
    if versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    profile_index = request.args.get("profile", default=0, type=int)
    if profile_index is None or profile_index < 0:
        return jsonify(error="Profilindex muss >= 0 sein."), 400

    try:
        analyze_path, error = _resolve_analyze_path_for_experiment(experiment_id)
        if error:
            status = 404 if error.startswith("Keine Analyse-Datei") else 400
            if error.startswith("Analyse-Datei nicht gefunden"):
                status = 404
            return jsonify(error=error), status

        graph_data = _load_scanning_profile_graph(analyze_path, profile_index)
        return jsonify(id=experiment_id, **graph_data)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    except Exception as exc:
        return jsonify(error=f"Fehler beim Lesen der Analyse-Datei: {exc}"), 500


@app.get("/api/analyzing/<experiment_id>/heatmap")
def get_analyzing_heatmap(experiment_id):
    if versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    try:
        analyze_path, error = _resolve_analyze_path_for_experiment(experiment_id)
        if error:
            status = 404 if error.startswith("Keine Analyse-Datei") else 400
            if error.startswith("Analyse-Datei nicht gefunden"):
                status = 404
            return jsonify(error=error), status

        heatmap_data = _load_analyzing_scan_heatmap(analyze_path)
        return jsonify(id=experiment_id, **heatmap_data)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    except Exception as exc:
        return jsonify(error=f"Fehler beim Lesen der Analyse-Heatmap: {exc}"), 500


@app.get("/api/analyzing/<experiment_id>/weld-graph")
def get_analyzing_weld_graph(experiment_id):
    if versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    try:
        analyze_weld_path, error = _resolve_analyze_weld_path_for_experiment(experiment_id)
        if error:
            status = 404 if error.startswith("Keine Analyse-H5-Datei") else 400
            if error.startswith("Analyse-H5-Datei nicht gefunden"):
                status = 404
            return jsonify(error=error), status

        analyze_weld_filename = analyze_weld_path.name
        graph_data = _load_welding_graph_from_h5(analyze_weld_path)
        return jsonify(id=experiment_id, h5file=analyze_weld_filename, **graph_data)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    except Exception as exc:
        return jsonify(error=f"Fehler beim Lesen der Analyse-H5-Datei: {exc}"), 500


@app.post("/api/load-versuchsuebersicht")
def load_versuchsuebersicht():
    body = request.get_json(silent=True) or {}
    path_str = body.get("path", "").strip()

    if not path_str:
        return jsonify(error="Bitte einen Pfad angeben."), 400

    path = _resolve_path(path_str)
    if not path.is_file():
        return jsonify(error=f"Datei nicht gefunden: {path_str}"), 404

    try:
        result = _load_versuchsuebersicht_from_path(path)
        return jsonify(success=True, **result)
    except Exception as exc:
        return jsonify(error=f"Fehler beim Laden: {exc}"), 500


@app.get("/api/versuchsuebersicht/status")
def versuchsuebersicht_status():
    status = _get_load_status()
    if versuchsuebersicht_data is None:
        return jsonify(loaded=False, **status)

    return jsonify(
        loaded=True,
        path=versuchsuebersicht_path,
        row_count=len(versuchsuebersicht_data["rows"]),
        column_count=len(versuchsuebersicht_data["columns"]),
        **status,
    )


@app.get("/api/welding/<experiment_id>")
def get_welding_data(experiment_id):
    if versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return jsonify(error="ID nicht gefunden."), 404

    field_names = list(
        dict.fromkeys(_flat_welding_param_columns() + WELDING_FORM_COLUMNS)
    )
    fields = _get_experiment_fields(row, field_names)

    is_welded = (
        _experiment_has_valid_weld(experiment_id)
        or _excel_row_has_weld_metadata(row)
    )

    return jsonify(id=experiment_id, fields=fields, is_welded=is_welded)


@app.post("/api/welding/reset/<experiment_id>")
def reset_welding(experiment_id):
    if versuchsuebersicht_path is None or versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return jsonify(error="ID nicht gefunden."), 404

    has_weld = (
        _experiment_has_valid_weld(experiment_id)
        or _excel_row_has_weld_metadata(row)
    )
    if not has_weld:
        return jsonify(error="Für diese ID ist kein Schweißvorgang gespeichert."), 400

    try:
        updated_fields = _reset_experiment_weld_in_excel(
            Path(versuchsuebersicht_path),
            experiment_id,
        )
        _load_versuchsuebersicht_from_path(Path(versuchsuebersicht_path))
        return jsonify(
            success=True,
            id=experiment_id,
            is_welded=False,
            fields={
                "WELDED": updated_fields["WELDED"],
                "H5FILE": updated_fields["H5FILE"],
            },
            deleted_file=updated_fields.get("deleted_file"),
        )
    except Exception as exc:
        return jsonify(error=f"Schweißen konnte nicht zurückgesetzt werden: {exc}"), 500


@app.post("/api/welding/start/<experiment_id>")
def start_welding(experiment_id):
    if versuchsuebersicht_path is None or versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return jsonify(error="ID nicht gefunden."), 404

    if (
        _experiment_has_valid_weld(experiment_id)
        or _excel_row_has_weld_metadata(row)
    ):
        return jsonify(error="Diese ID wurde bereits geschweißt."), 400

    existing_job = _get_weld_job(experiment_id)
    if existing_job and existing_job.get("status") == "running":
        return jsonify(error="Schweißen läuft bereits."), 409

    try:
        WELDING_DATA_DIR.mkdir(parents=True, exist_ok=True)
        _set_weld_job(experiment_id, {"status": "running"})
        thread = threading.Thread(
            target=_run_weld_job,
            args=(experiment_id, WELD_USE_DEMO),
            daemon=True,
        )
        thread.start()
        return jsonify(success=True, started=True, id=experiment_id, demo_mode=WELD_USE_DEMO)
    except Exception as exc:
        return jsonify(error=f"Schweißen konnte nicht gestartet werden: {exc}"), 500


@app.get("/api/welding/<experiment_id>/weld-status")
def get_welding_status(experiment_id):
    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    job = _get_weld_job(experiment_id)
    run_weld = _import_run_weld()
    progress = run_weld.get_weld_progress(experiment_id)

    if job and job.get("status") == "running":
        payload = {
            "running": True,
            "id": experiment_id,
            "status": progress.get("status") if progress else "running",
            "message": progress.get("message") if progress else "Schweißen läuft...",
            "elapsed_s": progress.get("elapsed_s") if progress else 0,
            "demo_mode": progress.get("demo_mode") if progress else WELD_USE_DEMO,
            "current_a": progress.get("current_a") if progress else None,
            "voltage_v": progress.get("voltage_v") if progress else None,
        }
        return jsonify(payload)

    if not job:
        return jsonify(running=False, id=experiment_id)

    if job.get("status") == "error":
        error = job.get("error") or "Schweißen fehlgeschlagen."
        _clear_weld_job(experiment_id)
        return jsonify(running=False, id=experiment_id, error=error)

    if job.get("status") == "done":
        payload = dict(job.get("result") or {})
        payload.update(
            success=True,
            running=False,
            id=experiment_id,
            is_welded=True,
            fields=job.get("fields") or {},
        )
        _clear_weld_job(experiment_id)
        return jsonify(payload)

    return jsonify(running=False, id=experiment_id)


@app.get("/api/welding/<experiment_id>/live-plot")
def get_welding_live_plot(experiment_id):
    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    since = request.args.get("since", "0")
    try:
        since_index = max(0, int(since))
    except ValueError:
        since_index = 0

    run_weld = _import_run_weld()
    plot_data = run_weld.get_weld_plot_data(experiment_id, since=since_index)
    if plot_data is not None:
        return jsonify(plot_data)

    progress = run_weld.get_weld_progress(experiment_id)
    if progress:
        return jsonify(
            active=True,
            points=[],
            total=0,
            current_a=progress.get("current_a"),
            voltage_v=progress.get("voltage_v"),
        )

    return jsonify(active=False, points=[], total=0, current_a=None, voltage_v=None)


@app.post("/api/scanning/start/<experiment_id>")
def start_scanning(experiment_id):
    if versuchsuebersicht_path is None or versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return jsonify(error="ID nicht gefunden."), 404

    if _experiment_has_valid_scan(experiment_id):
        return jsonify(error="Diese ID wurde bereits gescannt."), 400

    try:
        body = request.get_json(silent=True) or {}
        ui_settings = body.get("scan_settings")
        if ui_settings:
            scan_settings = _get_scan_settings_from_ui(ui_settings)
        else:
            scan_settings = _get_scan_settings_from_row(row)

        existing_job = _get_scan_job(experiment_id)
        if existing_job and existing_job.get("status") == "running":
            return jsonify(error="Scan läuft bereits."), 409

        resolved_scanspeed = _resolve_scanspeed_value(body.get("scanspeed"), row)

        thread = threading.Thread(
            target=_run_scan_job,
            args=(experiment_id, scan_settings, SCAN_USE_DEMO),
            kwargs={"scanspeed": resolved_scanspeed},
            daemon=True,
        )
        thread.start()
        return jsonify(success=True, started=True, id=experiment_id)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    except Exception as exc:
        return jsonify(error=f"Scan fehlgeschlagen: {exc}"), 500


@app.get("/api/scanning/<experiment_id>/scan-status")
def get_scanning_status(experiment_id):
    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    job = _get_scan_job(experiment_id)
    if not job:
        return jsonify(running=False, id=experiment_id)

    if job.get("status") == "running":
        return jsonify(running=True, id=experiment_id)

    if job.get("status") == "error":
        error = job.get("error") or "Scan fehlgeschlagen."
        _clear_scan_job(experiment_id)
        return jsonify(running=False, id=experiment_id, error=error)

    if job.get("status") == "done":
        payload = dict(job.get("result") or {})
        payload.update(
            success=True,
            running=False,
            id=experiment_id,
            is_scanned=True,
            fields=job.get("fields") or {},
            scan_settings=job.get("scan_settings"),
        )
        _clear_scan_job(experiment_id)
        return jsonify(payload)

    return jsonify(running=False, id=experiment_id)


@app.post("/api/scanning/reset/<experiment_id>")
def reset_scanning(experiment_id):
    if versuchsuebersicht_path is None or versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return jsonify(error="ID nicht gefunden."), 404

    has_scan = (
        _experiment_has_valid_scan(experiment_id)
        or _excel_row_has_scan_metadata(row)
    )
    if not has_scan:
        return jsonify(error="Für diese ID ist kein Scan gespeichert."), 400

    try:
        updated_fields = _reset_experiment_scan_in_excel(
            Path(versuchsuebersicht_path),
            experiment_id,
        )
        _load_versuchsuebersicht_from_path(Path(versuchsuebersicht_path))
        return jsonify(
            success=True,
            id=experiment_id,
            is_scanned=False,
            fields={
                "SCANNED": updated_fields["SCANNED"],
                "JSONFILE": updated_fields["JSONFILE"],
            },
            deleted_file=updated_fields.get("deleted_file"),
        )
    except Exception as exc:
        return jsonify(error=f"Scan konnte nicht zurückgesetzt werden: {exc}"), 500


@app.get("/api/scanning/<experiment_id>")
def get_scanning_data(experiment_id):
    if versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))
    if row is None:
        return jsonify(error="ID nicht gefunden."), 404

    field_names = list(
        dict.fromkeys(_flat_welding_param_columns() + ["PATH ID", "COMMENT"])
    )

    repaired_fields = _repair_orphaned_scan_metadata(experiment_id)
    if repaired_fields is not None:
        _load_versuchsuebersicht_from_path(Path(versuchsuebersicht_path))
        row = versuchsuebersicht_data["id_index"].get(_normalize_id(experiment_id))

    fields = _apply_scan_display_fields(
        _get_experiment_fields(row, field_names),
        experiment_id,
    )
    is_scanned = _experiment_has_valid_scan(experiment_id)

    return jsonify(id=experiment_id, fields=fields, is_scanned=is_scanned)


@app.get("/api/scanning/<experiment_id>/graph")
def get_scanning_graph(experiment_id):
    if versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    profile_index = request.args.get("profile", default=0, type=int)
    if profile_index is None or profile_index < 0:
        return jsonify(error="Profilindex muss >= 0 sein."), 400

    try:
        scan_path, error = _resolve_scan_path_for_experiment(experiment_id)
        if error:
            status = 404 if error.startswith("Keine Scan-Datei") else 400
            if error.startswith("Scan-Datei nicht gefunden"):
                status = 404
            return jsonify(error=error), status

        graph_data = _load_scanning_profile_graph(scan_path, profile_index)
        return jsonify(id=experiment_id, **graph_data)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    except Exception as exc:
        return jsonify(error=f"Fehler beim Lesen der Scan-Datei: {exc}"), 500


@app.get("/api/scanning/<experiment_id>/profile-y")
def get_scanning_profile_y(experiment_id):
    if versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    try:
        scan_path, error = _resolve_scan_path_for_experiment(experiment_id)
        if error:
            status = 404 if error.startswith("Keine Scan-Datei") else 400
            if error.startswith("Scan-Datei nicht gefunden"):
                status = 404
            return jsonify(error=error), status

        document = _load_scan_document(scan_path)
        profiles = document.get("profiles") or []
        if not profiles:
            return jsonify(error="Keine Profile in Scan-Datei gefunden."), 404

        y_mm = _get_profile_y_mm_values(
            document,
            profiles,
            experiment_id=document.get("experiment_id"),
        )
        return jsonify(
            id=experiment_id,
            profile_count=len(y_mm),
            y_mm=y_mm,
            y_min=min(y_mm),
            y_max=max(y_mm),
        )
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    except Exception as exc:
        return jsonify(error=f"Fehler beim Lesen der Profilpositionen: {exc}"), 500


@app.get("/api/scanning/<experiment_id>/live-profile")
def get_scanning_live_profile(experiment_id):
    if versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    try:
        job = _get_scan_job(experiment_id)
        if job and job.get("status") == "running":
            profile, demo_mode, uses_default_settings = _capture_live_profile_for_experiment(experiment_id)
            graph_data = _build_profile_graph_payload(
                profile,
                live=True,
                demo_mode=demo_mode,
                uses_default_settings=uses_default_settings,
                scanning=True,
            )
            return jsonify(id=experiment_id, **graph_data)

        if _experiment_has_valid_scan(experiment_id):
            return jsonify(error="Für diese ID liegt bereits ein Scan vor."), 409

        profile, demo_mode, uses_default_settings = _capture_live_profile_for_experiment(experiment_id)
        graph_data = _build_profile_graph_payload(
            profile,
            live=True,
            demo_mode=demo_mode,
            uses_default_settings=uses_default_settings,
        )
        return jsonify(id=experiment_id, **graph_data)
    except ValueError as exc:
        return jsonify(error=str(exc)), 503
    except Exception as exc:
        _close_live_profile_preview()
        message = str(exc)
        if "-303" in message:
            message = (
                "Scanner-Verbindung vorübergehend blockiert. "
                "Bitte 1–2 Sekunden warten und die Seite ggf. neu laden."
            )
        return jsonify(error=message), 500


@app.post("/api/scanning/<experiment_id>/live-profile/stop")
def stop_scanning_live_profile(experiment_id):
    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    _close_live_profile_preview()
    return jsonify(success=True, id=experiment_id)


@app.get("/api/welding/<experiment_id>/graph")
def get_welding_graph(experiment_id):
    if versuchsuebersicht_data is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    try:
        h5_filename = _get_experiment_h5_filename(experiment_id)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400

    if not h5_filename:
        return jsonify(error="Keine H5-Datei für diese ID."), 404

    h5_path = WELDING_DATA_DIR / h5_filename
    if not h5_path.is_file():
        return jsonify(error=f"H5-Datei nicht gefunden: {h5_filename}"), 404

    try:
        graph_data = _load_welding_graph_from_h5(h5_path)
        return jsonify(id=experiment_id, h5file=h5_filename, **graph_data)
    except Exception as exc:
        return jsonify(error=f"Fehler beim Lesen der H5-Datei: {exc}"), 500


@app.post("/api/welding/<experiment_id>")
def save_welding_data(experiment_id):
    if versuchsuebersicht_path is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    body = request.get_json(silent=True) or {}
    field_values = body.get("fields", {})

    try:
        _validate_welding_field_values(field_values)
        path = Path(versuchsuebersicht_path)
        updated_fields = _save_welding_fields_to_excel(path, experiment_id, field_values)
        _load_versuchsuebersicht_from_path(path)
        return jsonify(success=True, id=experiment_id, fields=updated_fields)
    except Exception as exc:
        return jsonify(error=f"Fehler beim Speichern: {exc}"), 500


@app.post("/api/scanning/<experiment_id>")
def save_scanning_data(experiment_id):
    if versuchsuebersicht_path is None:
        return jsonify(error="Keine Versuchsübersicht geladen. Bitte oben auf Laden klicken."), 400

    if not re.fullmatch(r"[A-Z]{3}", experiment_id):
        return jsonify(error="ID muss aus 3 Großbuchstaben bestehen."), 400

    body = request.get_json(silent=True) or {}
    field_values = body.get("fields", {})

    try:
        path = Path(versuchsuebersicht_path)
        _repair_orphaned_scan_metadata(experiment_id)
        is_scanned = _experiment_has_valid_scan(experiment_id)
        editable_columns = (
            SCANNING_EDITABLE_AFTER_SCAN_COLUMNS
            if is_scanned
            else SCANNING_EDITABLE_COLUMNS
        )
        updated_fields = _save_scanning_fields_to_excel(
            path,
            experiment_id,
            field_values,
            editable_columns=editable_columns,
        )
        _load_versuchsuebersicht_from_path(path)
        is_scanned = _experiment_has_valid_scan(experiment_id)
        return jsonify(
            success=True,
            id=experiment_id,
            is_scanned=is_scanned,
            fields=updated_fields,
        )
    except Exception as exc:
        return jsonify(error=f"Fehler beim Speichern: {exc}"), 500


@app.get("/path-images/<path_id>")
def path_image(path_id):
    normalized_path_id = path_id.strip().upper()
    if normalized_path_id not in PATH_ID_OPTIONS:
        return jsonify(error="PATH ID nicht gefunden."), 404

    return send_from_directory(str(PATH_IMAGES_DIR), f"{normalized_path_id}.svg")


@app.get("/styles.css")
def styles():
    return send_from_directory(TEMPLATES_DIR, "styles.css")


@app.get("/<path:path>")
def static_files(path):
    return send_from_directory(BASE_DIR, path)


_try_load_default_versuchsuebersicht()

if __name__ == "__main__":
    if SCAN_USE_DEMO:
        print("Scan-Demo-Modus aktiv (SCAN_USE_DEMO=1) – Scans ohne Hardware.")
    if WELD_USE_DEMO:
        print("Schweiß-Demo-Modus aktiv (WELD_USE_DEMO=1) – Schweißen ohne NI-DAQ.")
    app.run(debug=True, port=5000)
